# database/import_stories.py
# ============================================================================
# IMPORT STORIES FROM EXCEL TO DATABASE
# ============================================================================
# Purpose: Import refined user stories directly into the Client Product Database
#          without going through the full parsing/generation pipeline.
#
# Use Cases:
#   - Import manually refined user stories from a draft Excel
#   - Seed the database with existing work from another system
#   - Migrate stories from spreadsheets into the database
#   - Build reference library from historical approved stories
#
# AVIATION ANALOGY:
#   Think of this as "direct entry" into the flight plan system -
#   instead of going through the full flight planning process,
#   you're entering pre-approved waypoints directly.
#
# ============================================================================

import os
import re
from datetime import datetime
from typing import Optional, Dict, List, Any, Tuple

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================================
# FLEXIBLE COLUMN MAPPING
# ============================================================================
# WHY: Different Excel files may have different column naming conventions.
# This mapping allows us to match columns case-insensitively and with
# variations in naming.
#
# R EQUIVALENT:
#   column_mappings <- list(
#       story_id = c("story id", "id", "identifier"),
#       title = c("title", "name", "summary"),
#       ...
#   )

COLUMN_MAPPINGS = {
    # Core story fields
    'story_id': [
        'story id', 'id', 'identifier', 'story_id', 'storyid',
        'story-id', 'user story id', 'req id', 'requirement id'
    ],
    'title': [
        'title', 'name', 'summary', 'story title', 'feature',
        'feature title', 'requirement', 'functionality'
    ],
    'user_story': [
        'user story', 'story', 'description', 'user_story', 'as a',
        'story description', 'user story text', 'story text'
    ],
    'role': [
        'role', 'actor', 'user', 'persona', 'user role', 'as a role'
    ],
    'acceptance_criteria': [
        'acceptance criteria', 'ac', 'criteria', 'acceptance_criteria',
        'acceptance', 'acc criteria', 'acceptance crit'
    ],
    'success_metrics': [
        'success metrics', 'metrics', 'success_metrics', 'kpi',
        'success criteria', 'measurable outcomes'
    ],
    'priority': [
        'priority', 'ranking', 'importance', 'pri', 'prio',
        'priority level', 'urgency'
    ],
    'status': [
        'status', 'state', 'approval', 'approval status',
        'workflow status', 'review status'
    ],
    'category': [
        'category', 'type', 'cat', 'category code', 'feature type',
        'requirement type'
    ],
    'is_technical': [
        'is technical', 'technical', 'is_technical', 'tech',
        'technical item', 'is tech'
    ],

    # Notes and feedback fields
    'internal_notes': [
        'internal notes', 'your notes', 'notes', 'internal_notes',
        'team notes', 'reviewer notes'
    ],
    'meeting_context': [
        'meeting context', 'meeting notes', 'meeting_context',
        'discussion notes', 'meeting decisions'
    ],
    'client_feedback': [
        'client feedback', 'client notes', 'feedback', 'client_feedback',
        'customer feedback', 'stakeholder feedback'
    ],

    # Source tracking
    'source_requirement': [
        'source requirement', 'source', 'original requirement',
        'source_requirement', 'original text', 'source req'
    ],
    'source_row': [
        'source row', 'row', 'source_row', 'original row', 'row number'
    ]
}


def _normalize_header(header: str) -> str:
    """
    PURPOSE:
        Normalize a header string for comparison.
        Converts to lowercase, strips whitespace, removes special chars.

    PARAMETERS:
        header (str): Original header text

    RETURNS:
        str: Normalized header for matching

    R EQUIVALENT:
        normalize_header <- function(header) {
            tolower(trimws(gsub("[^a-z0-9 ]", "", header)))
        }
    """
    if not header:
        return ''

    # Convert to lowercase and strip whitespace
    normalized = str(header).lower().strip()

    # Remove special characters but keep spaces
    normalized = re.sub(r'[^a-z0-9\s]', '', normalized)

    # Collapse multiple spaces to single
    normalized = re.sub(r'\s+', ' ', normalized)

    return normalized


def _find_column_index(
    headers: List[str],
    field_name: str
) -> Optional[int]:
    """
    PURPOSE:
        Find the column index for a field using flexible matching.
        Tries all variations defined in COLUMN_MAPPINGS.

    PARAMETERS:
        headers (list): List of header strings from the Excel file
        field_name (str): The field we're looking for (e.g., 'story_id')

    RETURNS:
        int or None: Column index (0-based) if found, None otherwise

    WHY THIS APPROACH:
        Different Excel files use different column names. By checking
        multiple variations, we can handle files from various sources
        without requiring exact column names.
    """
    if field_name not in COLUMN_MAPPINGS:
        return None

    variations = COLUMN_MAPPINGS[field_name]

    # Normalize all headers once
    normalized_headers = [_normalize_header(h) for h in headers]

    # Try each variation
    for variation in variations:
        normalized_var = _normalize_header(variation)

        for idx, normalized_header in enumerate(normalized_headers):
            # Exact match
            if normalized_header == normalized_var:
                return idx

            # Partial match (header contains variation)
            if normalized_var in normalized_header:
                return idx

    return None


def _build_column_map(headers: List[str]) -> Dict[str, int]:
    """
    PURPOSE:
        Build a mapping from field names to column indices.

    PARAMETERS:
        headers (list): List of header strings from row 1

    RETURNS:
        dict: {field_name: column_index, ...}

    EXAMPLE:
        {'story_id': 0, 'title': 1, 'user_story': 2, ...}
    """
    column_map = {}

    for field_name in COLUMN_MAPPINGS.keys():
        idx = _find_column_index(headers, field_name)
        if idx is not None:
            column_map[field_name] = idx

    return column_map


def _extract_category_from_id(story_id: str) -> str:
    """
    PURPOSE:
        Extract category code from story ID (e.g., PROP-RECRUIT-001 → RECRUIT)

    PARAMETERS:
        story_id (str): Story ID like "PROP-RECRUIT-001"

    RETURNS:
        str: Category code or 'GEN' if not found
    """
    if not story_id:
        return 'GEN'

    # Pattern: PREFIX-CATEGORY-SEQ
    parts = str(story_id).split('-')
    if len(parts) >= 2:
        return parts[1].upper()

    return 'GEN'


def _parse_is_technical(value: Any) -> bool:
    """
    PURPOSE:
        Parse is_technical field from various formats.

    PARAMETERS:
        value: Could be bool, string 'Yes'/'No'/'True'/'False', 1/0

    RETURNS:
        bool: True if technical, False otherwise
    """
    if value is None:
        return True  # Default to technical

    if isinstance(value, bool):
        return value

    str_val = str(value).lower().strip()

    # Falsy values
    if str_val in ['no', 'false', '0', 'n', 'f', 'non-technical', 'workflow']:
        return False

    # Truthy values (or anything else defaults to True)
    return True


def _generate_story_id(prefix: str, category: str, sequence: int) -> str:
    """
    PURPOSE:
        Generate a story ID in the standard format.

    PARAMETERS:
        prefix (str): Program prefix (e.g., 'PROP')
        category (str): Category code (e.g., 'RECRUIT')
        sequence (int): Sequence number

    RETURNS:
        str: Story ID like "PROP-RECRUIT-001"
    """
    return f"{prefix.upper()}-{category.upper()}-{sequence:03d}"


def import_stories_from_excel(
    db_manager,
    excel_path: str,
    client_name: str,
    program_name: str,
    prefix: str,
    default_status: str = 'Approved',
    sheet_name: Optional[str] = None,
    add_to_reference: bool = False,
    verbose: bool = False
) -> Dict[str, Any]:
    """
    PURPOSE:
        Import refined user stories directly into database from Excel.
        Bypasses the full parsing/generation pipeline.

    Use this when:
        - You have manually refined user stories outside the pipeline
        - You want to seed the database with existing work
        - You're migrating stories from another system
        - You're building a reference library from historical approved stories

    AVIATION ANALOGY:
        This is "direct entry" into the flight management system -
        entering pre-approved waypoints directly without going through
        the full route planning workflow.

    PARAMETERS:
        db_manager: ClientProductDatabase instance
        excel_path (str): Path to Excel file with user stories
        client_name (str): Client name (creates if not exists)
        program_name (str): Program name (creates if not exists)
        prefix (str): Program prefix (e.g., 'PROP')
        default_status (str): Status for stories without explicit status
                             Default: 'Approved' (assumes refined = approved)
        sheet_name (str, optional): Specific sheet to read
        add_to_reference (bool): Add approved stories to reference library
        verbose (bool): Print detailed progress

    EXPECTED EXCEL COLUMNS (flexible matching):
        - Story ID (or ID, Identifier) - optional, auto-generated if missing
        - Title (or Name, Summary) - REQUIRED
        - User Story (or Story, Description)
        - Role (or Actor, User)
        - Acceptance Criteria (or AC, Criteria)
        - Success Metrics (or Metrics)
        - Priority (or Ranking, Importance)
        - Status (or State) - defaults to 'Approved' if not present
        - Category (optional) - extracted from ID or defaults to 'GEN'
        - Internal Notes / Your Notes (optional)
        - Meeting Context (optional)
        - Client Feedback (optional)
        - Is Technical (optional) - defaults to True

    PROCESS:
        1. Create or find client
        2. Create or find program by prefix
        3. Read Excel file
        4. Map columns flexibly (handle different column names)
        5. For each row:
           - Generate story_id if not present (PREFIX-CAT-SEQ)
           - Set status to 'Approved' if not specified
           - Insert into user_stories table
           - Log to audit_history
           - Optionally add to reference library
        6. Return summary

    RETURNS:
        {
            'success': True,
            'imported': 15,
            'updated': 2,
            'skipped': 1,
            'errors': ['Row 5: Missing title'],
            'client_id': 'xxx',
            'program_id': 'xxx',
            'story_ids': ['PROP-RECRUIT-001', ...]
        }

    R EQUIVALENT:
        import_stories_from_excel <- function(db, excel_path, client, program, prefix) {
            # Read Excel
            stories <- readxl::read_excel(excel_path)

            # Match columns flexibly
            col_map <- match_columns(names(stories))

            # Insert each row
            for (row in 1:nrow(stories)) {
                story <- build_story_dict(stories[row, ], col_map)
                insert_story(db, story)
            }
        }
    """
    # ========================================================================
    # VALIDATION
    # ========================================================================
    if not OPENPYXL_AVAILABLE:
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': ['openpyxl not installed. Run: pip install openpyxl'],
            'client_id': None,
            'program_id': None,
            'story_ids': []
        }

    if not os.path.exists(excel_path):
        return {
            'success': False,
            'imported': 0,
            'updated': 0,
            'skipped': 0,
            'errors': [f'File not found: {excel_path}'],
            'client_id': None,
            'program_id': None,
            'story_ids': []
        }

    # Initialize result
    result = {
        'success': False,
        'imported': 0,
        'updated': 0,
        'skipped': 0,
        'errors': [],
        'client_id': None,
        'program_id': None,
        'story_ids': []
    }

    # ========================================================================
    # DATABASE SETUP
    # ========================================================================
    try:
        # Find or create client
        existing_client = db_manager.get_client_by_name(client_name)
        if existing_client:
            client_id = existing_client['client_id']
            if verbose:
                print(f"  → Found existing client: {client_name}")
        else:
            client_id = db_manager.create_client(client_name)
            if verbose:
                print(f"  ✓ Created client: {client_name} ({client_id})")

        result['client_id'] = client_id

        # Find or create program by prefix
        existing_program = db_manager.get_program_by_prefix(prefix)
        if existing_program:
            program_id = existing_program['program_id']
            if verbose:
                print(f"  → Found existing program: {existing_program['name']} ({prefix})")
        else:
            program_id = db_manager.create_program(
                client_id=client_id,
                name=program_name,
                prefix=prefix,
                source_file=os.path.basename(excel_path)
            )
            if verbose:
                print(f"  ✓ Created program: {program_name} ({prefix})")

        result['program_id'] = program_id

    except Exception as e:
        result['errors'].append(f"Database setup failed: {e}")
        return result

    # ========================================================================
    # READ EXCEL FILE
    # ========================================================================
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)

        # Select sheet
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                result['errors'].append(f"Sheet not found: {sheet_name}")
                return result
            ws = wb[sheet_name]
        else:
            # Use first sheet
            ws = wb.active

        if verbose:
            print(f"  → Reading sheet: {ws.title}")

    except Exception as e:
        result['errors'].append(f"Failed to open Excel file: {e}")
        return result

    # ========================================================================
    # MAP COLUMNS
    # ========================================================================
    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else '')

    column_map = _build_column_map(headers)

    if verbose:
        print(f"  → Mapped {len(column_map)} columns:")
        for field, idx in sorted(column_map.items()):
            print(f"      • {field}: column {idx + 1} ({headers[idx]})")

    # Validate required columns
    if 'title' not in column_map:
        result['errors'].append("Required column 'Title' not found in Excel")
        return result

    # ========================================================================
    # PROCESS ROWS
    # ========================================================================
    # Track sequences by category for auto-generated IDs
    category_sequences = {}

    # Get existing story IDs to detect updates
    existing_stories = {
        s['story_id']: s
        for s in db_manager.get_stories(program_id, status_filter=None)
    }

    stories_to_save = []
    row_num = 1  # Header is row 1

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_num += 1

        # Skip empty rows
        if not any(row):
            continue

        # Helper to get cell value by field name
        def get_value(field_name: str, default: Any = None) -> Any:
            if field_name not in column_map:
                return default
            idx = column_map[field_name]
            if idx < len(row):
                val = row[idx]
                return val if val is not None else default
            return default

        # ----------------------------------------------------------------
        # Extract core fields
        # ----------------------------------------------------------------
        title = get_value('title', '')
        if not title or not str(title).strip():
            result['errors'].append(f"Row {row_num}: Missing title (skipped)")
            result['skipped'] += 1
            continue

        title = str(title).strip()

        # Story ID (auto-generate if not present)
        story_id = get_value('story_id', '')
        category = get_value('category', '')

        if not story_id:
            # Auto-generate story ID
            if not category:
                category = 'GEN'
            else:
                category = str(category).upper()[:10]

            # Get next sequence for this category
            if category not in category_sequences:
                # Count existing stories in this category
                existing_count = sum(
                    1 for sid in existing_stories.keys()
                    if sid and f"-{category}-" in sid
                )
                category_sequences[category] = existing_count + 1
            else:
                category_sequences[category] += 1

            story_id = _generate_story_id(prefix, category, category_sequences[category])

        else:
            story_id = str(story_id).strip()
            # Extract category from ID if not explicitly provided
            if not category:
                category = _extract_category_from_id(story_id)

        # ----------------------------------------------------------------
        # Extract other fields
        # ----------------------------------------------------------------
        user_story = get_value('user_story', '')
        role = get_value('role', 'user')
        acceptance_criteria = get_value('acceptance_criteria', '')
        success_metrics = get_value('success_metrics', '')
        priority = get_value('priority', 'Medium')
        status = get_value('status', default_status)
        is_technical = _parse_is_technical(get_value('is_technical', True))
        internal_notes = get_value('internal_notes', '')
        meeting_context = get_value('meeting_context', '')
        client_feedback = get_value('client_feedback', '')
        source_requirement = get_value('source_requirement', '')
        source_row = get_value('source_row', row_num)

        # Normalize status
        status = str(status).strip() if status else default_status
        valid_statuses = ['Draft', 'Internal Review', 'Pending Client Review',
                         'Approved', 'Needs Discussion', 'Out of Scope']
        if status not in valid_statuses:
            # Try case-insensitive match
            status_lower = status.lower()
            matched = False
            for vs in valid_statuses:
                if vs.lower() == status_lower:
                    status = vs
                    matched = True
                    break
            if not matched:
                status = default_status

        # Normalize priority
        priority = str(priority).strip() if priority else 'Medium'
        valid_priorities = ['Critical', 'High', 'Medium', 'Low']
        if priority not in valid_priorities:
            priority_lower = priority.lower()
            for vp in valid_priorities:
                if vp.lower() == priority_lower:
                    priority = vp
                    break

        # ----------------------------------------------------------------
        # Build story dict
        # ----------------------------------------------------------------
        story = {
            'story_id': story_id,
            'generated_id': story_id,  # Alias for compatibility
            'title': title,
            'user_story': str(user_story).strip() if user_story else '',
            'role': str(role).strip() if role else 'user',
            'acceptance_criteria': str(acceptance_criteria).strip() if acceptance_criteria else '',
            'success_metrics': str(success_metrics).strip() if success_metrics else '',
            'priority': priority,
            'status': status,
            'category_abbrev': category,
            'category': category,
            'is_technical': is_technical,
            'internal_notes': str(internal_notes).strip() if internal_notes else '',
            'meeting_context': str(meeting_context).strip() if meeting_context else '',
            'client_feedback': str(client_feedback).strip() if client_feedback else '',
            'source_requirement': {
                'description': str(source_requirement).strip() if source_requirement else '',
                'row_number': source_row
            },
            'flags': []  # No flags for imported stories
        }

        # Check if this is an update or insert
        is_update = story_id in existing_stories

        stories_to_save.append(story)
        result['story_ids'].append(story_id)

        if verbose:
            action = "Update" if is_update else "Import"
            print(f"      {action}: {story_id} - {title[:40]}...")

    # ========================================================================
    # SAVE TO DATABASE
    # ========================================================================
    if stories_to_save:
        try:
            inserted, updated = db_manager.save_user_stories(program_id, stories_to_save)
            result['imported'] = inserted
            result['updated'] = updated

            if verbose:
                print(f"  ✓ Saved {inserted} new, {updated} updated stories")

            # Add approved stories to reference library
            if add_to_reference:
                ref_count = 0
                for story in stories_to_save:
                    if story['status'] == 'Approved':
                        try:
                            keywords = f"{story['category']},{story['title']}"
                            db_manager.add_to_reference_library(
                                story['story_id'],
                                keywords,
                                quality_score=4  # Manually imported = good quality
                            )
                            ref_count += 1
                        except Exception:
                            pass  # Ignore reference library errors

                if verbose and ref_count > 0:
                    print(f"  ✓ Added {ref_count} stories to reference library")

        except Exception as e:
            result['errors'].append(f"Database save failed: {e}")
            return result

    # ========================================================================
    # SUCCESS
    # ========================================================================
    result['success'] = len(result['errors']) == 0

    return result


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def quick_import(
    excel_path: str,
    prefix: str,
    client_name: str = "Default Client",
    program_name: Optional[str] = None
) -> Dict[str, Any]:
    """
    PURPOSE:
        Quick import function that handles database connection.

    PARAMETERS:
        excel_path (str): Path to Excel file
        prefix (str): Program prefix
        client_name (str): Client name (default: "Default Client")
        program_name (str): Program name (default: "{PREFIX} Program")

    RETURNS:
        dict: Import result

    USAGE:
        from database.import_stories import quick_import
        result = quick_import("stories.xlsx", "PROP", "Acme Corp")
    """
    from .db_manager import get_database

    db = get_database()
    prog_name = program_name or f"{prefix} Program"

    result = import_stories_from_excel(
        db_manager=db,
        excel_path=excel_path,
        client_name=client_name,
        program_name=prog_name,
        prefix=prefix,
        verbose=True
    )

    db.close()
    return result


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    import sys

    print("=" * 70)
    print("IMPORT STORIES TEST")
    print("=" * 70)

    # Check for test file argument
    if len(sys.argv) < 2:
        print("\nUsage: python3 import_stories.py <excel_file> [prefix]")
        print("\nExample:")
        print("  python3 import_stories.py outputs/drafts/PROP_draft_user_stories.xlsx PROP")
        print("\nThis will import stories from the Excel file into the test database.")
        sys.exit(1)

    excel_path = sys.argv[1]
    prefix = sys.argv[2] if len(sys.argv) > 2 else "TEST"

    print(f"\nImporting from: {excel_path}")
    print(f"Prefix: {prefix}")
    print()

    # Use test database
    from .db_manager import ClientProductDatabase
    import os

    test_db_path = "data/test_import_stories.db"
    if os.path.exists(test_db_path):
        os.remove(test_db_path)

    db = ClientProductDatabase(test_db_path)

    result = import_stories_from_excel(
        db_manager=db,
        excel_path=excel_path,
        client_name="Test Client",
        program_name="Test Program",
        prefix=prefix,
        default_status="Approved",
        verbose=True
    )

    print("\n" + "=" * 70)
    print("IMPORT RESULT")
    print("=" * 70)
    print(f"Success: {result['success']}")
    print(f"Imported: {result['imported']}")
    print(f"Updated: {result['updated']}")
    print(f"Skipped: {result['skipped']}")
    print(f"Errors: {len(result['errors'])}")

    if result['errors']:
        print("\nErrors:")
        for err in result['errors']:
            print(f"  • {err}")

    if result['story_ids']:
        print(f"\nStory IDs ({len(result['story_ids'])}):")
        for sid in result['story_ids'][:10]:
            print(f"  • {sid}")
        if len(result['story_ids']) > 10:
            print(f"  ... and {len(result['story_ids']) - 10} more")

    db.close()
    print(f"\nTest database: {test_db_path}")
