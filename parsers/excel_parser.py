# parsers/excel_parser.py
# ============================================================================
# PURPOSE: Parse Excel requirements documents with FLEXIBLE column handling
#
# DESIGN PHILOSOPHY:
#     Don't assume column names — capture EVERYTHING from the spreadsheet.
#     Map "core" columns by flexible pattern matching, store everything else
#     as context_columns for downstream use.
#
# AVIATION ANALOGY:
#     Like a flight data recorder that captures ALL parameters, not just the
#     ones you think are important. You never know what context will matter
#     when analyzing the requirement later.
#
# R EQUIVALENT:
#     Like read_excel() %>% rename_with(~detect_semantic_field(.x)) but
#     keeping all original columns as additional context.
#
# ============================================================================

import re
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell


class ExcelParser:
    """
    PURPOSE:
        Parse Excel files containing requirements into normalized dictionaries.
        Captures ALL columns from the spreadsheet:
        - "Core" columns are mapped to standard fields (title, description, priority, etc.)
        - All other columns are stored in context_columns dict

    WHY THIS APPROACH:
        Real-world requirements documents have varying column names and include
        stakeholder notes, decisions, and clarifications in arbitrary columns.
        By capturing everything, we give downstream generators full context.

    USAGE:
        parser = ExcelParser("inputs/excel/requirements.xlsx")
        requirements = parser.parse()  # Excludes completed by default
        requirements = parser.parse(include_completed=True)  # Include completed

    OUTPUT STRUCTURE:
        {
            "requirement_id": "REQ-001" or None,
            "title": "...",
            "description": "...",
            "priority": "High",
            "status": "Planned",
            "type": "Technical Feature",
            "row_number": 5,
            "source_file": "requirements.xlsx",
            "context_columns": {
                "Notes": "...",
                "Supplemental Notes": "...",
                "Impact": "...",
                "Whatever Else": "..."
            }
        }
    """

    # ========================================================================
    # CORE COLUMN PATTERNS
    # ========================================================================
    # These patterns identify "core" columns that map to standard fields.
    # All patterns are case-insensitive and use partial matching.
    # Order matters: first match wins.

    CORE_FIELD_PATTERNS = {
        # Primary content fields
        'title': [
            r'\btitle\b', r'\bname\b', r'\bfunctionality\b', r'\bfeature\b',
            r'\bitem\b', r'\bsubject\b'
        ],
        'description': [
            r'\bdescription\b', r'\bdetail\b', r'\bspec\b', r'\brequirement\b',
            r'\bstory\b', r'\bnarrative\b'
        ],

        # Priority/importance
        'priority': [
            r'\bpriority\b', r'\bpri\b', r'\branking\b', r'\brank\b',
            r'\bimportance\b', r'\bmoscow\b', r'\burgency\b'
        ],

        # Status field - used for filtering
        'status': [
            r'\bstatus\b', r'\bstate\b', r'\bprogress\b', r'\bphase\b'
        ],

        # Type/category
        'type': [
            r'\btype\b', r'\bcategory\b', r'\bclassification\b', r'\bmodule\b',
            r'\barea\b', r'\bgroup\b', r'\btheme\b'
        ],

        # ID fields
        'requirement_id': [
            r'\bid\b', r'\breq\s*#', r'\bitem\s*#', r'\bref\b', r'\b#\b'
        ],
    }

    # Status values that indicate "completed" (case-insensitive)
    COMPLETED_STATUS_VALUES = [
        'completed', 'complete', 'done', 'finished', 'closed',
        'released', 'shipped', 'deployed', 'implemented'
    ]

    def __init__(
        self,
        file_path: str,
        include_completed: bool = False
    ) -> None:
        """
        PURPOSE:
            Initialize the parser with file path and configuration.

        PARAMETERS:
            file_path (str): Path to the Excel file
            include_completed (bool): Whether to include completed requirements
                                     Default: False (skip completed items)

        WHY include_completed DEFAULT FALSE:
            Most users want to work on active requirements. Completed items
            add noise and don't need new test cases. Users can override
            with include_completed=True when needed.
        """
        self.file_path = Path(file_path)
        self.include_completed = include_completed
        self.workbook = None

        # Will be populated during parsing
        self._header_row: Optional[int] = None
        self._headers: dict[int, str] = {}  # col_index -> original header name
        self._core_mapping: dict[str, int] = {}  # field_name -> col_index
        self._context_columns: list[int] = []  # col indices not mapped to core

        # Stats tracking
        self.stats = {
            'total_rows': 0,
            'requirements_found': 0,
            'completed_skipped': 0,
            'empty_skipped': 0,
            'columns_found': 0,
            'core_columns_mapped': 0,
            'context_columns_found': 0
        }

    def parse(
        self,
        sheet_name: Optional[str] = None,
        include_completed: Optional[bool] = None
    ) -> list[dict]:
        """
        PURPOSE:
            Parse the Excel file and extract all requirements.

        PARAMETERS:
            sheet_name (Optional[str]): Sheet to parse (default: first sheet)
            include_completed (Optional[bool]): Override include_completed setting

        RETURNS:
            list[dict]: List of requirement dictionaries with structure:
                {
                    "requirement_id": str or None,
                    "title": str or None,
                    "description": str or None,
                    "priority": str (normalized),
                    "status": str or None,
                    "type": str or None,
                    "row_number": int,
                    "source_file": str,
                    "raw_text": str,  # For backward compatibility
                    "context_columns": dict  # All non-core columns
                }

        WHY THIS STRUCTURE:
            - Core fields are directly accessible for common operations
            - context_columns preserves ALL other data for downstream use
            - raw_text maintains backward compatibility with existing code
        """
        # Allow runtime override of include_completed
        if include_completed is not None:
            self.include_completed = include_completed

        # Load workbook
        self._load_workbook()
        worksheet = self._get_worksheet(sheet_name)

        # Detect headers and build column mapping
        self._detect_and_map_headers(worksheet)

        # Extract requirements
        requirements = self._extract_all_requirements(worksheet)

        return requirements

    def _load_workbook(self) -> None:
        """Load the Excel file into memory."""
        if not self.file_path.exists():
            raise FileNotFoundError(f"Excel file not found: {self.file_path}")

        valid_extensions = ['.xlsx', '.xlsm', '.xls']
        if self.file_path.suffix.lower() not in valid_extensions:
            raise ValueError(f"Not an Excel file: {self.file_path.suffix}")

        try:
            self.workbook = openpyxl.load_workbook(
                self.file_path,
                data_only=True
            )
        except Exception as e:
            raise RuntimeError(f"Failed to open Excel file: {e}")

    def _get_worksheet(self, sheet_name: Optional[str] = None) -> Worksheet:
        """Get the target worksheet."""
        if self.workbook is None:
            raise RuntimeError("Workbook not loaded")

        if sheet_name is None:
            return self.workbook.active

        if sheet_name not in self.workbook.sheetnames:
            available = ", ".join(self.workbook.sheetnames)
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {available}")

        return self.workbook[sheet_name]

    def _detect_and_map_headers(self, worksheet: Worksheet) -> None:
        """
        PURPOSE:
            Find header row and map ALL columns:
            - Core columns → mapped to standard field names
            - Other columns → stored as context_columns

        WHY CAPTURE EVERYTHING:
            Context columns often contain the REAL requirement after stakeholder
            discussion. Notes, Decisions, Supplemental Notes, etc. contain
            critical information that should inform story generation.
        """
        # Find the header row (first row with 2+ header-like values)
        for row_idx in range(1, min(11, worksheet.max_row + 1)):
            row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx))[0]

            headers_in_row = {}
            for col_idx, cell in enumerate(row):
                value = self._get_cell_value(cell)
                if value and self._looks_like_header(value):
                    headers_in_row[col_idx] = value

            if len(headers_in_row) >= 2:
                self._header_row = row_idx
                self._headers = headers_in_row
                break

        if not self._headers:
            # No clear headers - use row 1 column names as-is
            row = list(worksheet.iter_rows(min_row=1, max_row=1))[0]
            for col_idx, cell in enumerate(row):
                value = self._get_cell_value(cell)
                if value:
                    self._headers[col_idx] = value
            self._header_row = 1

        # Now map core columns and identify context columns
        self._build_column_mapping()

        # Update stats
        self.stats['columns_found'] = len(self._headers)
        self.stats['core_columns_mapped'] = len(self._core_mapping)
        self.stats['context_columns_found'] = len(self._context_columns)

        # Log what we found
        print(f"[ExcelParser] Found {len(self._headers)} columns:")
        print(f"  Core columns mapped: {len(self._core_mapping)}")
        for field, col_idx in sorted(self._core_mapping.items()):
            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
            header = self._headers.get(col_idx, 'unknown')
            print(f"    - {field}: column {col_letter} ({header})")
        print(f"  Context columns: {len(self._context_columns)}")
        for col_idx in self._context_columns[:5]:  # Show first 5
            header = self._headers.get(col_idx, 'unknown')
            print(f"    - {header}")
        if len(self._context_columns) > 5:
            print(f"    ... and {len(self._context_columns) - 5} more")

    def _build_column_mapping(self) -> None:
        """
        PURPOSE:
            Map columns to core fields and identify context columns.

        APPROACH:
            1. Try to match each header to core field patterns
            2. First match wins (prevents double-mapping)
            3. Unmatched columns become context_columns
        """
        mapped_columns = set()

        # Map core fields
        for field_name, patterns in self.CORE_FIELD_PATTERNS.items():
            for col_idx, header_text in self._headers.items():
                if col_idx in mapped_columns:
                    continue  # Already mapped to another field

                header_lower = header_text.lower()
                for pattern in patterns:
                    if re.search(pattern, header_lower, re.IGNORECASE):
                        self._core_mapping[field_name] = col_idx
                        mapped_columns.add(col_idx)
                        break

                if field_name in self._core_mapping:
                    break  # Found mapping for this field

        # All unmapped columns are context columns
        self._context_columns = [
            col_idx for col_idx in self._headers.keys()
            if col_idx not in mapped_columns
        ]

    def _looks_like_header(self, text: str) -> bool:
        """Check if text looks like a column header."""
        if not text or len(text) > 100:
            return False

        # Headers are typically short and don't contain requirement language
        text_lower = text.lower()

        # Requirement indicators suggest this is content, not a header
        requirement_words = ['shall', 'must', 'should', 'as a', 'i want']
        if any(word in text_lower for word in requirement_words):
            return False

        # Very long text is probably content
        if len(text) > 60:
            return False

        return True

    def _extract_all_requirements(self, worksheet: Worksheet) -> list[dict]:
        """
        PURPOSE:
            Extract requirements from all data rows.

        APPROACH:
            - Skip header row
            - Skip empty rows
            - Skip completed rows (unless include_completed=True)
            - Extract core fields + all context columns
        """
        requirements = []
        self.stats['total_rows'] = worksheet.max_row

        for row_idx, row in enumerate(worksheet.iter_rows(), start=1):
            # Skip header row
            if row_idx == self._header_row:
                continue

            # Extract the row data
            result = self._extract_row(list(row), row_idx)

            if result is None:
                self.stats['empty_skipped'] += 1
                continue

            # Check if this is a completed item
            if self._is_completed(result) and not self.include_completed:
                self.stats['completed_skipped'] += 1
                continue

            requirements.append(result)
            self.stats['requirements_found'] += 1

        return requirements

    def _extract_row(self, cells: list[Cell], row_idx: int) -> Optional[dict]:
        """
        PURPOSE:
            Extract a single row into a requirement dictionary.

        RETURNS:
            dict with core fields + context_columns, or None if empty row
        """
        # Helper to get cell value by column index
        def get_cell(col_idx: int) -> Optional[str]:
            if col_idx < len(cells):
                return self._get_cell_value(cells[col_idx])
            return None

        # Extract core fields
        title = get_cell(self._core_mapping.get('title', -1))
        description = get_cell(self._core_mapping.get('description', -1))
        priority_raw = get_cell(self._core_mapping.get('priority', -1))
        status = get_cell(self._core_mapping.get('status', -1))
        req_type = get_cell(self._core_mapping.get('type', -1))
        req_id = get_cell(self._core_mapping.get('requirement_id', -1))

        # Must have at least title or description
        if not title and not description:
            return None

        # Skip very short content (likely empty or just whitespace)
        content = (description or title or "").strip()
        if len(content) < 3:
            return None

        # Extract ALL context columns
        context_columns = {}
        for col_idx in self._context_columns:
            header = self._headers.get(col_idx, f'Column_{col_idx}')
            value = get_cell(col_idx)
            if value:  # Only include non-empty values
                context_columns[header] = value

        # Also include core columns that have extra info in context
        # (e.g., if there's an "Impact" column not mapped to core)
        for col_idx, header in self._headers.items():
            if col_idx not in self._core_mapping.values():
                continue
            # Check if this is a field we didn't map
            value = get_cell(col_idx)
            if value and header not in context_columns:
                # Don't duplicate core fields in context
                pass

        # Normalize priority
        priority = self._normalize_priority(priority_raw)

        # Build requirement dict
        requirement = {
            # Core fields
            'requirement_id': req_id,
            'title': title,
            'description': description,
            'priority': priority,
            'status': status,
            'type': req_type,

            # Traceability
            'row_number': row_idx,
            'source_file': str(self.file_path.name),

            # All context columns (notes, decisions, impacts, etc.)
            'context_columns': context_columns,

            # Backward compatibility
            'raw_text': description or title,
        }

        return requirement

    def _is_completed(self, requirement: dict) -> bool:
        """Check if a requirement is completed based on status."""
        status = requirement.get('status', '')
        if not status:
            return False

        status_lower = status.lower().strip()
        return any(
            completed in status_lower
            for completed in self.COMPLETED_STATUS_VALUES
        )

    def _normalize_priority(self, priority_text: Optional[str]) -> str:
        """
        PURPOSE:
            Normalize priority to standard values: Critical, High, Medium, Low

        RETURNS:
            str: Normalized priority (defaults to "Medium" if unrecognized)
        """
        if not priority_text:
            return "Medium"

        priority_lower = priority_text.lower().strip()

        # Priority mapping
        priority_map = {
            # Critical
            'critical': 'Critical', 'crit': 'Critical', 'urgent': 'Critical',
            'must have': 'Critical', 'must': 'Critical', 'p0': 'Critical',
            'mo': 'Critical',
            # High
            'high': 'High', 'h': 'High', 'important': 'High',
            'should have': 'High', 'should': 'High', 'p1': 'High',
            'sc': 'High',
            # Medium
            'medium': 'Medium', 'med': 'Medium', 'm': 'Medium',
            'normal': 'Medium', 'could have': 'Medium', 'could': 'Medium',
            'p2': 'Medium', 'co': 'Medium',
            # Low
            'low': 'Low', 'l': 'Low', 'minor': 'Low',
            'nice to have': 'Low', 'would': 'Low', 'p3': 'Low',
            'wo': 'Low',
        }

        return priority_map.get(priority_lower, priority_text.title())

    def _get_cell_value(self, cell: Cell) -> Optional[str]:
        """Safely extract string value from a cell."""
        if isinstance(cell, MergedCell):
            return None
        if cell.value is None:
            return None

        value = str(cell.value).strip()
        return value if value else None

    def get_stats(self) -> dict:
        """Return parsing statistics."""
        return self.stats.copy()


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def parse_excel(
    file_path: str,
    sheet_name: Optional[str] = None,
    include_completed: bool = False
) -> list[dict]:
    """
    PURPOSE:
        Convenience function to parse an Excel file.

    PARAMETERS:
        file_path (str): Path to Excel file
        sheet_name (Optional[str]): Sheet to parse (default: first)
        include_completed (bool): Include completed requirements (default: False)

    RETURNS:
        list[dict]: List of requirement dictionaries

    USAGE:
        requirements = parse_excel("inputs/excel/requirements.xlsx")
    """
    parser = ExcelParser(file_path, include_completed=include_completed)
    return parser.parse(sheet_name=sheet_name)


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python excel_parser.py <path_to_excel_file> [--include-completed]")
        sys.exit(1)

    file_path = sys.argv[1]
    include_completed = "--include-completed" in sys.argv

    print(f"Parsing: {file_path}")
    print(f"Include completed: {include_completed}")
    print("-" * 60)

    parser = ExcelParser(file_path, include_completed=include_completed)
    requirements = parser.parse()

    print(f"\n{'='*60}")
    print(f"RESULTS: Found {len(requirements)} requirements")
    print(f"{'='*60}")

    stats = parser.get_stats()
    print(f"\nStats:")
    print(f"  Total rows: {stats['total_rows']}")
    print(f"  Requirements found: {stats['requirements_found']}")
    print(f"  Completed skipped: {stats['completed_skipped']}")
    print(f"  Empty skipped: {stats['empty_skipped']}")

    # Show first 3 requirements in detail
    for i, req in enumerate(requirements[:3], start=1):
        print(f"\n{'-'*60}")
        print(f"Requirement {i} (Row {req['row_number']})")
        print(f"{'-'*60}")

        if req.get('requirement_id'):
            print(f"  ID: {req['requirement_id']}")
        if req.get('title'):
            print(f"  Title: {req['title'][:80]}...")
        if req.get('description'):
            desc = req['description']
            print(f"  Description: {desc[:100]}{'...' if len(desc) > 100 else ''}")
        if req.get('priority'):
            print(f"  Priority: {req['priority']}")
        if req.get('status'):
            print(f"  Status: {req['status']}")
        if req.get('type'):
            print(f"  Type: {req['type']}")

        # Show context columns
        context = req.get('context_columns', {})
        if context:
            print(f"  Context columns ({len(context)}):")
            for key, value in list(context.items())[:5]:
                val_preview = str(value)[:60] + '...' if len(str(value)) > 60 else value
                print(f"    - {key}: {val_preview}")
            if len(context) > 5:
                print(f"    ... and {len(context) - 5} more")
