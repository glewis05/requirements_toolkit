# parsers/user_story_parser.py
# ============================================================================
# PURPOSE: Import refined user stories from Excel (Phase 2 of two-phase workflow)
#
# This parser reads user stories that have been reviewed and refined by
# stakeholders, then passes them to the UAT generator for test case creation.
#
# TWO-PHASE WORKFLOW:
#   Phase 1: Requirements → Draft Stories → Human Review
#   Phase 2: Refined Stories (this parser) → UAT Test Cases
#
# EXPECTED INPUT FORMAT (from draft_excel_formatter.py):
#   Sheet: "User Stories for Review"
#   Columns: Story ID, Title, User Story, Role, Acceptance Criteria,
#            Success Metrics, Priority, Status, Your Notes, Meeting Context,
#            Is Technical, Source Requirement, Source Row
#
# AVIATION ANALOGY:
#   Like importing a reviewed and signed flight plan. The pilot has added
#   their notes, weather considerations, and fuel calculations. Now we
#   execute based on the final plan, not the original draft.
#
# ============================================================================

import os
from typing import Optional

# openpyxl for Excel file reading
try:
    from openpyxl import load_workbook
except ImportError:
    raise ImportError("openpyxl required. Install with: pip3 install openpyxl")


class UserStoryParser:
    """
    PURPOSE:
        Parse refined user stories from Excel for UAT generation.

    KEY FEATURES:
        - Reads stories from "User Stories for Review" sheet
        - Filters by status (Approved, Draft vs Out of Scope)
        - Flags stories with "Needs Discussion" status
        - Extracts reviewer notes and meeting context
        - Preserves traceability to source requirements

    OUTPUT STRUCTURE:
        {
            "story_id": "PROP-RECRUIT-001",
            "title": "...",
            "user_story": "As a...",
            "role": "research coordinator",
            "acceptance_criteria": ["...", "..."],
            "success_metrics": ["...", "..."],
            "priority": "High",
            "status": "Approved",
            "reviewer_notes": "...",
            "meeting_context": "...",
            "client_feedback": "...",  # NEW: Client signoff comments
            "is_technical": True,
            "source_requirement": "...",
            "source_row": 5,
            "flags": ["needs_discussion"]  # if applicable
        }

    USAGE:
        parser = UserStoryParser("refined_stories.xlsx")
        stories = parser.parse()
    """

    # Expected column headers and their indices (updated for 14 columns)
    EXPECTED_COLUMNS = {
        'Story ID': 0,
        'Title': 1,
        'User Story': 2,
        'Role': 3,
        'Acceptance Criteria': 4,
        'Success Metrics': 5,
        'Priority': 6,
        'Status': 7,
        'Your Notes': 8,
        'Meeting Context': 9,
        'Client Feedback': 10,  # NEW: Client signoff comments
        'Is Technical': 11,
        'Source Requirement': 12,
        'Source Row': 13,
    }

    def __init__(self, filepath: str, sheet_name: Optional[str] = None):
        """
        PURPOSE:
            Initialize the parser with file path.

        PARAMETERS:
            filepath (str): Path to the refined stories Excel file
            sheet_name (str, optional): Sheet name to parse
                                       Default: "User Stories for Review"
        """
        self.filepath = filepath
        self.sheet_name = sheet_name or "User Stories for Review"

        # Validate file exists
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")

        # Statistics
        self.stats = {
            'total_rows': 0,
            'stories_included': 0,
            'stories_excluded': 0,
            'approved': 0,
            'draft': 0,
            'pending_client_review': 0,
            'needs_discussion': 0,
            'out_of_scope': 0,
            'technical': 0,
            'non_technical': 0,
        }

    def parse(self) -> list[dict]:
        """
        PURPOSE:
            Parse refined user stories from Excel.

        RETURNS:
            list[dict]: List of user story dictionaries

        FILTER LOGIC:
            - Include: Status = "Approved", "Draft", or "Pending Client Review"
            - Exclude: Status = "Out of Scope"
            - Flag: Status = "Needs Discussion" (included but flagged)
        """
        wb = load_workbook(self.filepath, data_only=True)

        # Find the sheet
        if self.sheet_name not in wb.sheetnames:
            # Try to find a similar sheet
            available = ', '.join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{self.sheet_name}' not found.\n"
                f"Available sheets: {available}"
            )

        ws = wb[self.sheet_name]

        # Detect column mapping from header row
        column_map = self._detect_columns(ws)

        # Parse stories
        stories = []

        for row_idx in range(2, ws.max_row + 1):
            story = self._parse_row(ws, row_idx, column_map)

            if story is None:
                continue

            self.stats['total_rows'] += 1

            # Filter by status
            status = story.get('status', 'Draft')

            if status == 'Out of Scope':
                self.stats['stories_excluded'] += 1
                self.stats['out_of_scope'] += 1
                continue

            # Include the story
            self.stats['stories_included'] += 1

            if status == 'Approved':
                self.stats['approved'] += 1
            elif status == 'Draft':
                self.stats['draft'] += 1
            elif status == 'Pending Client Review':
                self.stats['pending_client_review'] += 1
                story['flags'].append('pending_client_review')
            elif status == 'Needs Discussion':
                self.stats['needs_discussion'] += 1
                story['flags'].append('needs_discussion')

            # Track technical vs non-technical
            if story.get('is_technical', True):
                self.stats['technical'] += 1
            else:
                self.stats['non_technical'] += 1

            stories.append(story)

        return stories

    def _detect_columns(self, ws) -> dict:
        """
        PURPOSE:
            Detect column indices from header row.

        RETURNS:
            dict: Maps column name to column index (1-based)
        """
        column_map = {}
        header_row = 1

        for col_idx in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col_idx).value
            if header:
                header = str(header).strip()
                column_map[header] = col_idx

        return column_map

    def _parse_row(self, ws, row_idx: int, column_map: dict) -> Optional[dict]:
        """
        PURPOSE:
            Parse a single row into a user story dict.

        RETURNS:
            dict or None if row is empty
        """
        def get_cell(col_name: str, default=''):
            """Helper to get cell value by column name."""
            col_idx = column_map.get(col_name)
            if col_idx:
                value = ws.cell(row=row_idx, column=col_idx).value
                return value if value is not None else default
            return default

        # Get story ID - if empty, skip row
        story_id = get_cell('Story ID', '')
        if not story_id or str(story_id).strip() == '':
            return None

        # Parse acceptance criteria (multi-line string to list)
        ac_text = get_cell('Acceptance Criteria', '')
        acceptance_criteria = self._parse_multiline(ac_text)

        # Parse success metrics (multi-line string to list)
        metrics_text = get_cell('Success Metrics', '')
        success_metrics = self._parse_multiline(metrics_text)

        # Parse is_technical
        is_tech_str = str(get_cell('Is Technical', 'Yes')).strip().lower()
        is_technical = is_tech_str in ['yes', 'true', '1', 'y']

        # Parse source row
        source_row = get_cell('Source Row', '')
        try:
            source_row = int(source_row) if source_row else None
        except (ValueError, TypeError):
            source_row = None

        # Build source_requirement as dict for traceability compatibility
        source_req_text = str(get_cell('Source Requirement', '')).strip()
        source_requirement = {
            'requirement_id': str(story_id).strip(),  # Use story ID as req ID
            'description': source_req_text,
            'row_number': source_row,
        }

        return {
            'story_id': str(story_id).strip(),
            'generated_id': str(story_id).strip(),  # Alias for compatibility
            'title': str(get_cell('Title', '')).strip(),
            'user_story': str(get_cell('User Story', '')).strip(),
            'role': str(get_cell('Role', 'user')).strip(),
            'acceptance_criteria': acceptance_criteria,
            'success_metrics': success_metrics,
            'priority': str(get_cell('Priority', 'Medium')).strip(),
            'status': str(get_cell('Status', 'Draft')).strip(),
            'reviewer_notes': str(get_cell('Your Notes', '')).strip(),
            'meeting_context': str(get_cell('Meeting Context', '')).strip(),
            'client_feedback': str(get_cell('Client Feedback', '')).strip(),  # Client signoff
            'is_technical': is_technical,
            'source_requirement': source_requirement,  # Dict for traceability
            'source_requirement_text': source_req_text,  # Keep text version too
            'source_row': source_row,
            'flags': [],
            # Add fields for compatibility with UAT generator
            'capability': self._extract_capability(str(get_cell('User Story', ''))),
            'category_abbrev': self._extract_category(str(story_id)),
        }

    def _parse_multiline(self, text: str) -> list[str]:
        """
        PURPOSE:
            Parse multi-line text into a list of lines.
        """
        if not text:
            return []

        lines = str(text).split('\n')
        return [line.strip() for line in lines if line.strip()]

    def _extract_capability(self, user_story: str) -> str:
        """
        PURPOSE:
            Extract capability from user story text.
            "As a X, I want to Y, so that Z" → Y
        """
        import re
        match = re.search(r'I want(?:\s+to)?\s+(.+?)(?:,\s*so that|$)', user_story, re.IGNORECASE)
        if match:
            return match.group(1).strip()
        return user_story[:50]  # Fallback to first 50 chars

    def _extract_category(self, story_id: str) -> str:
        """
        PURPOSE:
            Extract category from story ID.
            "PROP-RECRUIT-001" → "RECRUIT"
        """
        parts = story_id.split('-')
        if len(parts) >= 2:
            return parts[1]
        return 'GEN'

    def get_stats(self) -> dict:
        """Return parsing statistics."""
        return self.stats.copy()


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def parse_refined_stories(
    filepath: str,
    sheet_name: Optional[str] = None
) -> list[dict]:
    """
    PURPOSE:
        Convenience function to parse refined user stories.

    PARAMETERS:
        filepath (str): Path to refined stories Excel file
        sheet_name (str, optional): Sheet name to parse

    RETURNS:
        list[dict]: List of user story dictionaries
    """
    parser = UserStoryParser(filepath, sheet_name)
    return parser.parse()


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    import sys

    print("=" * 70)
    print("USER STORY PARSER TEST")
    print("=" * 70)

    if len(sys.argv) > 1:
        filepath = sys.argv[1]
    else:
        print("\nUsage: python3 parsers/user_story_parser.py <refined_stories.xlsx>")
        print("\nTo test, first create a draft file:")
        print("  python3 run.py \"requirements.xlsx\" --prefix TEST --phase draft")
        print("  Then parse the generated draft file.")
        sys.exit(0)

    try:
        parser = UserStoryParser(filepath)
        stories = parser.parse()

        print(f"\nParsed {len(stories)} stories from: {filepath}")
        print(f"\nStatistics:")
        stats = parser.get_stats()
        for key, value in stats.items():
            print(f"  {key}: {value}")

        print(f"\nSample stories:")
        for story in stories[:3]:
            print(f"\n  {'-'*60}")
            print(f"  ID: {story['story_id']}")
            print(f"  Title: {story['title']}")
            print(f"  Status: {story['status']}")
            print(f"  Technical: {story['is_technical']}")
            print(f"  Acceptance Criteria: {len(story['acceptance_criteria'])} items")
            if story['flags']:
                print(f"  Flags: {', '.join(story['flags'])}")

    except Exception as e:
        print(f"\nERROR: {e}")
        sys.exit(1)
