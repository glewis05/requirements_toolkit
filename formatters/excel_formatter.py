# formatters/excel_formatter.py
# ============================================================================
# Excel Formatter for UAT Packages
# ============================================================================
#
# PURPOSE:
#     Exports user stories and UAT test cases to Excel format, compatible
#     with UAT packages like the GenoRx template. Creates a structured
#     workbook with multiple sheets for test cases, stories, and summary.
#
# AVIATION ANALOGY:
#     Like creating a flight test card package. You have the master test
#     matrix (all tests), individual procedure sheets (stories), and a
#     summary cover sheet. Each formatted for the audience who needs it.
#
# R EQUIVALENT:
#     In R, you'd use the openxlsx package:
#     library(openxlsx)
#     wb <- createWorkbook()
#     addWorksheet(wb, "Test Cases")
#     writeData(wb, "Test Cases", data)
#     saveWorkbook(wb, "output.xlsx")
#
# USAGE:
#     # Quick function call:
#     from formatters.excel_formatter import export_to_excel
#     filepath = export_to_excel(stories, test_cases, filename="uat_package.xlsx")
#
#     # Or use the class for more control:
#     from formatters.excel_formatter import ExcelFormatter
#     formatter = ExcelFormatter(output_dir='outputs/excel')
#     filepath = formatter.export(stories, test_cases)
#
# ============================================================================

import os
from datetime import datetime
from typing import Optional

# openpyxl for Excel file creation
# WHY: openpyxl is the standard Python library for .xlsx files
# R EQUIVALENT: Like the openxlsx package in R
from openpyxl import Workbook
from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)
from openpyxl.utils import get_column_letter


class ExcelFormatter:
    """
    PURPOSE:
        Export user stories and UAT test cases to Excel format.
        Creates a professionally formatted workbook suitable for
        UAT packages and stakeholder review.

    R EQUIVALENT:
        # In R with openxlsx:
        # ExcelFormatter <- R6Class("ExcelFormatter",
        #     public = list(
        #         output_dir = NULL,
        #         initialize = function(output_dir = "outputs/excel") {
        #             self$output_dir <- output_dir
        #         },
        #         export = function(stories, tests) {
        #             wb <- createWorkbook()
        #             # ... add sheets ...
        #             saveWorkbook(wb, file.path(self$output_dir, "output.xlsx"))
        #         }
        #     )
        # )

    ATTRIBUTES:
        output_dir (str): Directory where Excel files are saved
        source_file (str): Name of source file (for documentation)

    WHY THIS CLASS:
        Encapsulating in a class allows:
        1. Consistent styling across multiple exports
        2. Reusable configuration
        3. Easier extension for new sheet types
    """

    def __init__(
        self,
        output_dir: str = "outputs/excel",
        source_file: Optional[str] = None
    ):
        """
        PURPOSE:
            Initialize the Excel formatter with configuration.

        R EQUIVALENT:
            # In R:
            # formatter <- list(
            #     output_dir = "outputs/excel",
            #     source_file = NULL
            # )

        PARAMETERS:
            output_dir (str): Directory to save Excel files.
                Default: "outputs/excel"
                Will be created if it doesn't exist.

            source_file (str, optional): Name of the source file.
                Used in documentation and summary sheet.

        RETURNS:
            None (constructor)

        WHY THIS APPROACH:
            Separating configuration from export allows reuse of the
            formatter with consistent settings across multiple exports.
        """
        # Store output directory
        # WHY: All Excel files go here, keeping outputs organized
        self.output_dir = output_dir

        # Store source file name for documentation
        # WHY: Traceability — knowing where data came from
        self.source_file = source_file or "Unknown"

        # Ensure output directory exists
        # WHY: Avoid errors when saving files
        # R EQUIVALENT: dir.create(output_dir, recursive = TRUE, showWarnings = FALSE)
        os.makedirs(self.output_dir, exist_ok=True)

        # ================================================================
        # STYLE DEFINITIONS
        # ================================================================
        # Define reusable styles for consistent formatting
        # WHY: Centralized styles make it easy to change look-and-feel

        # Header style — bold white text on dark blue background
        # R EQUIVALENT: createStyle(fontColour = "white", fgFill = "#4472C4", textDecoration = "bold")
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(
            start_color="4472C4",
            end_color="4472C4",
            fill_type="solid"
        )
        self.header_alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # Priority fills — color coding for priority levels
        # WHY: Visual priority indicators help with quick scanning
        self.priority_fills = {
            'Critical': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'Must Have': PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid"),
            'High': PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
            'Should Have': PatternFill(start_color="FFA94D", end_color="FFA94D", fill_type="solid"),
            'Medium': PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid"),
            'Could Have': PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid"),
            'Low': PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid"),
            'Won\'t Have': PatternFill(start_color="CED4DA", end_color="CED4DA", fill_type="solid")
        }

        # Data cell style — left aligned with wrap
        self.data_alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        )

        # Border style — thin borders for all cells
        thin_side = Side(style='thin', color='CCCCCC')
        self.cell_border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side
        )

        # Column width configurations
        # WHY: Pre-defined widths ensure readable output without manual adjustment
        self.test_case_column_widths = {
            'A': 15,   # Test ID
            'B': 15,   # Category
            'C': 40,   # Title
            'D': 35,   # Pre-Requisites
            'E': 50,   # Test Steps
            'F': 45,   # Expected Results
            'G': 12,   # MoSCoW
            'H': 10,   # Est. Time
            'I': 30,   # Notes
            'J': 12    # Test Type (bonus column)
        }

        self.story_column_widths = {
            'A': 15,   # Story ID
            'B': 35,   # Title
            'C': 50,   # User Story
            'D': 12,   # Priority
            'E': 15,   # Role
            'F': 40,   # Acceptance Criteria
            'G': 20,   # Flags
            'H': 10    # Source Row
        }

        self.rtm_column_widths = {
            'A': 15,   # Req ID
            'B': 45,   # Requirement
            'C': 15,   # Story ID
            'D': 35,   # Story Title
            'E': 40,   # Test Cases
            'F': 20,   # Compliance
            'G': 12    # Status
        }

        # Coverage status fills for RTM
        self.coverage_fills = {
            'Full': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),     # Green
            'Partial': PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid"),  # Yellow
            'None': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")      # Red
        }

    def export(
        self,
        user_stories: list[dict],
        test_cases: list[dict],
        filename: str = "uat_package.xlsx",
        traceability_matrix: dict = None
    ) -> str:
        """
        PURPOSE:
            Main entry point — export stories and tests to Excel workbook.

        R EQUIVALENT:
            # export <- function(stories, tests, filename = "uat_package.xlsx") {
            #     wb <- createWorkbook()
            #     addWorksheet(wb, "Summary")
            #     addWorksheet(wb, "Test Case Master")
            #     addWorksheet(wb, "User Stories")
            #     # ... write data ...
            #     saveWorkbook(wb, filename)
            # }

        PARAMETERS:
            user_stories (list[dict]): List of user story dicts from UserStoryGenerator.
                Each should have: generated_id, title, user_story, priority, etc.

            test_cases (list[dict]): List of test case dicts from UATGenerator.
                Each should have: test_id, title, test_steps, expected_results, etc.

            filename (str): Output filename.
                Default: "uat_package.xlsx"

            traceability_matrix (dict, optional): RTM from TraceabilityGenerator.
                If provided, adds a Traceability Matrix sheet.

        RETURNS:
            str: Full path to the created Excel file

        WHY THIS APPROACH:
            Creating a complete workbook with multiple sheets provides:
            1. Test Case Master for QA execution
            2. User Stories for context
            3. Summary for management overview
            4. Traceability Matrix for compliance/audit
        """
        # Create a new workbook
        # WHY: Each export gets a fresh workbook
        # R EQUIVALENT: wb <- createWorkbook()
        wb = Workbook()

        # Remove the default sheet created by openpyxl
        # WHY: We'll create our own sheets with specific names
        default_sheet = wb.active
        wb.remove(default_sheet)

        # ================================================================
        # CREATE SHEETS
        # ================================================================

        # 1. Summary sheet (first for visibility)
        self._create_summary_sheet(wb, user_stories, test_cases)

        # 2. Test Case Master sheet
        self._create_test_case_sheet(wb, test_cases)

        # 3. User Stories sheet
        self._create_user_stories_sheet(wb, user_stories)

        # 4. Traceability Matrix sheet (if provided)
        if traceability_matrix:
            self._create_traceability_sheet(wb, traceability_matrix)

        # ================================================================
        # SAVE WORKBOOK
        # ================================================================
        filepath = os.path.join(self.output_dir, filename)
        wb.save(filepath)

        return filepath

    def _create_summary_sheet(
        self,
        wb: Workbook,
        user_stories: list[dict],
        test_cases: list[dict]
    ):
        """
        PURPOSE:
            Create the Summary sheet with counts and flagged items.

        PARAMETERS:
            wb (Workbook): The workbook to add the sheet to
            user_stories (list[dict]): All user stories
            test_cases (list[dict]): All test cases

        WHY THIS APPROACH:
            Summary sheet provides quick overview for stakeholders.
            Placed first so it's visible when file is opened.
        """
        # Create the sheet
        # R EQUIVALENT: addWorksheet(wb, "Summary")
        ws = wb.create_sheet("Summary")

        # ================================================================
        # DOCUMENT HEADER
        # ================================================================
        ws['A1'] = "UAT Test Package Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:D1')

        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A4'] = f"Source: {self.source_file}"

        # ================================================================
        # OVERVIEW COUNTS
        # ================================================================
        ws['A6'] = "Overview"
        ws['A6'].font = Font(bold=True, size=14)

        ws['A7'] = "Total User Stories:"
        ws['B7'] = len(user_stories)
        ws['A8'] = "Total Test Cases:"
        ws['B8'] = len(test_cases)

        # ================================================================
        # PRIORITY BREAKDOWN
        # ================================================================
        ws['A10'] = "Stories by Priority"
        ws['A10'].font = Font(bold=True, size=14)

        priority_counts = {'Critical': 0, 'High': 0, 'Medium': 0, 'Low': 0}
        for story in user_stories:
            priority = story.get('priority', 'Medium')
            if priority in priority_counts:
                priority_counts[priority] += 1

        row = 11
        for priority, count in priority_counts.items():
            ws[f'A{row}'] = priority
            ws[f'B{row}'] = count
            if priority in self.priority_fills:
                ws[f'A{row}'].fill = self.priority_fills[priority]
            row += 1

        # ================================================================
        # TEST TYPE BREAKDOWN
        # ================================================================
        ws['A16'] = "Test Cases by Type"
        ws['A16'].font = Font(bold=True, size=14)

        test_type_counts = {
            'Happy Path': 0,
            'Negative': 0,
            'Edge Case': 0,
            'Boundary': 0
        }
        type_mapping = {
            'happy_path': 'Happy Path',
            'negative': 'Negative',
            'edge_case': 'Edge Case',
            'boundary': 'Boundary'
        }
        for tc in test_cases:
            test_type = tc.get('test_type', 'unknown')
            display_type = type_mapping.get(test_type, test_type)
            if display_type in test_type_counts:
                test_type_counts[display_type] += 1

        row = 17
        for test_type, count in test_type_counts.items():
            ws[f'A{row}'] = test_type
            ws[f'B{row}'] = count
            row += 1

        # ================================================================
        # MOSCOW BREAKDOWN
        # ================================================================
        ws['A22'] = "Test Cases by MoSCoW"
        ws['A22'].font = Font(bold=True, size=14)

        moscow_counts = {
            'Must Have': 0,
            'Should Have': 0,
            'Could Have': 0,
            'Won\'t Have': 0
        }
        for tc in test_cases:
            moscow = tc.get('moscow', 'Should Have')
            if moscow in moscow_counts:
                moscow_counts[moscow] += 1

        row = 23
        for moscow, count in moscow_counts.items():
            ws[f'A{row}'] = moscow
            ws[f'B{row}'] = count
            if moscow in self.priority_fills:
                ws[f'A{row}'].fill = self.priority_fills[moscow]
            row += 1

        # ================================================================
        # FLAGGED ITEMS
        # ================================================================
        flagged_stories = [s for s in user_stories if s.get('flags')]

        ws['A28'] = "Items Requiring Attention"
        ws['A28'].font = Font(bold=True, size=14)

        if flagged_stories:
            row = 29
            for story in flagged_stories:
                title = story.get('title', 'Untitled')[:50]
                flags = ', '.join(story.get('flags', []))
                ws[f'A{row}'] = title
                ws[f'B{row}'] = flags
                ws[f'A{row}'].fill = PatternFill(
                    start_color="FFF3CD",
                    end_color="FFF3CD",
                    fill_type="solid"
                )
                row += 1
        else:
            ws['A29'] = "No flagged items - all stories passed quality checks"
            ws['A29'].font = Font(italic=True, color="666666")

        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 40
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20

    def _create_test_case_sheet(
        self,
        wb: Workbook,
        test_cases: list[dict]
    ):
        """
        PURPOSE:
            Create the Test Case Master sheet with all UAT test cases.

        PARAMETERS:
            wb (Workbook): The workbook to add the sheet to
            test_cases (list[dict]): All test cases

        WHY THIS APPROACH:
            This is the main sheet QA teams will use for execution.
            Formatted for easy reading and printing.
        """
        # Create the sheet
        ws = wb.create_sheet("Test Case Master")

        # ================================================================
        # HEADER ROW
        # ================================================================
        headers = [
            'Test ID',
            'Category',
            'Title',
            'Pre-Requisites',
            'Test Steps',
            'Expected Results',
            'MoSCoW',
            'Est. Time',
            'Notes',
            'Test Type'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

        # ================================================================
        # DATA ROWS
        # ================================================================
        for row_idx, tc in enumerate(test_cases, 2):
            # Test ID
            ws.cell(row=row_idx, column=1, value=tc.get('test_id', 'N/A'))

            # Category
            ws.cell(row=row_idx, column=2, value=tc.get('category', 'General'))

            # Title
            ws.cell(row=row_idx, column=3, value=tc.get('title', 'Untitled'))

            # Pre-Requisites — join list with newlines
            prereqs = tc.get('prerequisites', [])
            prereqs_text = '\n'.join([f"• {p}" for p in prereqs])
            ws.cell(row=row_idx, column=4, value=prereqs_text)

            # Test Steps — already numbered, join with newlines
            steps = tc.get('test_steps', [])
            steps_text = '\n'.join(steps)
            ws.cell(row=row_idx, column=5, value=steps_text)

            # Expected Results — join with newlines
            expected = tc.get('expected_results', [])
            expected_text = '\n'.join(expected)
            ws.cell(row=row_idx, column=6, value=expected_text)

            # MoSCoW — with color coding
            moscow = tc.get('moscow', 'Should Have')
            cell = ws.cell(row=row_idx, column=7, value=moscow)
            if moscow in self.priority_fills:
                cell.fill = self.priority_fills[moscow]

            # Est. Time
            ws.cell(row=row_idx, column=8, value=tc.get('est_time', '5 min'))

            # Notes
            ws.cell(row=row_idx, column=9, value=tc.get('notes', ''))

            # Test Type — with readable format
            test_type = tc.get('test_type', 'unknown')
            type_display = {
                'happy_path': 'Happy Path',
                'negative': 'Negative',
                'edge_case': 'Edge Case',
                'boundary': 'Boundary'
            }.get(test_type, test_type)
            ws.cell(row=row_idx, column=10, value=type_display)

            # Apply data formatting to all cells in this row
            for col in range(1, 11):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = self.data_alignment
                cell.border = self.cell_border

        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        for col_letter, width in self.test_case_column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ================================================================
        # FREEZE HEADER ROW
        # ================================================================
        # WHY: Keeps headers visible when scrolling through many test cases
        # R EQUIVALENT: freezePane(wb, "Test Case Master", firstRow = TRUE)
        ws.freeze_panes = 'A2'

        # ================================================================
        # AUTO-FILTER
        # ================================================================
        # WHY: Allows easy filtering by category, type, or priority
        ws.auto_filter.ref = ws.dimensions

    def _create_user_stories_sheet(
        self,
        wb: Workbook,
        user_stories: list[dict]
    ):
        """
        PURPOSE:
            Create the User Stories sheet with all generated stories.

        PARAMETERS:
            wb (Workbook): The workbook to add the sheet to
            user_stories (list[dict]): All user stories

        WHY THIS APPROACH:
            Provides context for the test cases. QA can reference
            the original story when executing tests.
        """
        # Create the sheet
        ws = wb.create_sheet("User Stories")

        # ================================================================
        # HEADER ROW
        # ================================================================
        headers = [
            'Story ID',
            'Title',
            'User Story',
            'Priority',
            'Role',
            'Acceptance Criteria',
            'Quality Flags',
            'Source Row'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

        # ================================================================
        # DATA ROWS
        # ================================================================
        for row_idx, story in enumerate(user_stories, 2):
            # Story ID
            ws.cell(row=row_idx, column=1, value=story.get('generated_id', 'N/A'))

            # Title
            ws.cell(row=row_idx, column=2, value=story.get('title', 'Untitled'))

            # User Story (full format)
            ws.cell(row=row_idx, column=3, value=story.get('user_story', 'N/A'))

            # Priority — with color coding
            priority = story.get('priority', 'Medium')
            cell = ws.cell(row=row_idx, column=4, value=priority)
            if priority in self.priority_fills:
                cell.fill = self.priority_fills[priority]

            # Role
            ws.cell(row=row_idx, column=5, value=story.get('role', 'user'))

            # Acceptance Criteria — join with newlines
            criteria = story.get('acceptance_criteria', [])
            # Clean up bullet prefixes
            clean_criteria = []
            for c in criteria:
                clean = c.strip()
                if clean.startswith('•'):
                    clean = clean[1:].strip()
                clean_criteria.append(f"• {clean}")
            criteria_text = '\n'.join(clean_criteria)
            ws.cell(row=row_idx, column=6, value=criteria_text)

            # Quality Flags
            flags = story.get('flags', [])
            flags_text = ', '.join(flags) if flags else 'None'
            cell = ws.cell(row=row_idx, column=7, value=flags_text)
            if flags:
                cell.fill = PatternFill(
                    start_color="FFF3CD",
                    end_color="FFF3CD",
                    fill_type="solid"
                )

            # Source Row
            source_row = story.get('source_requirement', {}).get('row_number', 'N/A')
            ws.cell(row=row_idx, column=8, value=source_row)

            # Apply data formatting to all cells in this row
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = self.data_alignment
                cell.border = self.cell_border

        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        for col_letter, width in self.story_column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ================================================================
        # FREEZE HEADER ROW
        # ================================================================
        ws.freeze_panes = 'A2'

        # ================================================================
        # AUTO-FILTER
        # ================================================================
        ws.auto_filter.ref = ws.dimensions

    def _create_traceability_sheet(
        self,
        wb: Workbook,
        rtm: dict
    ):
        """
        PURPOSE:
            Create the Traceability Matrix sheet showing requirement-to-test mapping.

        PARAMETERS:
            wb (Workbook): The workbook to add the sheet to
            rtm (dict): Traceability matrix from TraceabilityGenerator

        WHY THIS APPROACH:
            Traceability matrices are essential for:
            - Regulatory compliance (Part 11, HIPAA, SOC 2)
            - Audit readiness
            - Gap analysis
            - Impact assessment when requirements change
        """
        ws = wb.create_sheet("Traceability Matrix")

        # ================================================================
        # COVERAGE SUMMARY SECTION
        # ================================================================
        ws['A1'] = "Requirements Traceability Matrix"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:G1')

        summary = rtm.get('summary', {})

        ws['A3'] = "Coverage Summary"
        ws['A3'].font = Font(bold=True, size=14)

        # Coverage stats
        ws['A4'] = "Total Requirements:"
        ws['B4'] = summary.get('total_requirements', 0)
        ws['C4'] = "Total Test Cases:"
        ws['D4'] = summary.get('total_test_cases', 0)

        ws['A5'] = "Full Coverage:"
        full_count = summary.get('full_coverage_count', 0)
        full_pct = summary.get('full_coverage_pct', 0)
        ws['B5'] = f"{full_count} ({full_pct}%)"
        ws['B5'].fill = self.coverage_fills['Full']

        ws['C5'] = "Partial Coverage:"
        partial_count = summary.get('partial_coverage_count', 0)
        partial_pct = summary.get('partial_coverage_pct', 0)
        ws['D5'] = f"{partial_count} ({partial_pct}%)"
        ws['D5'].fill = self.coverage_fills['Partial']

        ws['E5'] = "No Coverage:"
        none_count = summary.get('no_coverage_count', 0)
        none_pct = summary.get('no_coverage_pct', 0)
        ws['F5'] = f"{none_count} ({none_pct}%)"
        ws['F5'].fill = self.coverage_fills['None']

        # Compliance coverage
        ws['A7'] = "Compliance Test Coverage"
        ws['A7'].font = Font(bold=True, size=12)

        compliance = summary.get('compliance_coverage', {})
        col = 1
        for framework, stats in compliance.items():
            ws.cell(row=8, column=col, value=f"{framework}:")
            ws.cell(row=8, column=col + 1, value=f"{stats.get('tests', 0)} tests")
            col += 2

        # ================================================================
        # TRACEABILITY MATRIX TABLE
        # ================================================================
        ws['A10'] = "Detailed Traceability"
        ws['A10'].font = Font(bold=True, size=14)

        # Header row
        headers = [
            'Req ID',
            'Requirement',
            'Story ID',
            'Story Title',
            'Test Cases',
            'Compliance',
            'Status'
        ]

        header_row = 11
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.cell_border

        # Data rows
        matrix = rtm.get('matrix', [])
        for row_idx, row_data in enumerate(matrix, header_row + 1):
            # Req ID
            ws.cell(row=row_idx, column=1, value=row_data.get('requirement_id', ''))

            # Requirement (truncated)
            req_text = row_data.get('requirement_text', '')
            if len(req_text) > 100:
                req_text = req_text[:97] + '...'
            ws.cell(row=row_idx, column=2, value=req_text)

            # Story ID
            ws.cell(row=row_idx, column=3, value=row_data.get('user_story_id', ''))

            # Story Title
            story_title = row_data.get('user_story_title', '')
            if len(story_title) > 50:
                story_title = story_title[:47] + '...'
            ws.cell(row=row_idx, column=4, value=story_title)

            # Test Cases - show count and first few IDs
            test_ids = row_data.get('test_case_ids', [])
            test_count = row_data.get('test_case_count', 0)
            if test_ids:
                # Show first 3 test IDs, then count
                display_ids = test_ids[:3]
                if len(test_ids) > 3:
                    test_display = ', '.join(display_ids) + f' (+{len(test_ids) - 3} more)'
                else:
                    test_display = ', '.join(display_ids)
            else:
                test_display = 'None'
            ws.cell(row=row_idx, column=5, value=test_display)

            # Compliance frameworks
            compliance_coverage = row_data.get('compliance_coverage', [])
            compliance_display = ', '.join(compliance_coverage) if compliance_coverage else 'None'
            ws.cell(row=row_idx, column=6, value=compliance_display)

            # Status - with color coding
            status = row_data.get('coverage_status', 'None')
            cell = ws.cell(row=row_idx, column=7, value=status)
            if status in self.coverage_fills:
                cell.fill = self.coverage_fills[status]

            # Apply row fill based on status (entire row gets subtle color)
            for col in range(1, 8):
                cell = ws.cell(row=row_idx, column=col)
                cell.alignment = self.data_alignment
                cell.border = self.cell_border
                # Apply subtle row coloring
                if status == 'None':
                    cell.fill = PatternFill(start_color="FFEEEE", end_color="FFEEEE", fill_type="solid")
                elif status == 'Partial':
                    cell.fill = PatternFill(start_color="FFFBEE", end_color="FFFBEE", fill_type="solid")

        # ================================================================
        # GAPS SECTION
        # ================================================================
        gaps = rtm.get('gaps', [])
        if gaps:
            gaps_start_row = header_row + len(matrix) + 3

            ws.cell(row=gaps_start_row, column=1, value="Identified Gaps")
            ws.cell(row=gaps_start_row, column=1).font = Font(bold=True, size=14)

            for idx, gap in enumerate(gaps, gaps_start_row + 1):
                ws.cell(row=idx, column=1, value=gap.get('requirement_id', ''))
                ws.cell(row=idx, column=2, value=', '.join(gap.get('gaps', [])))
                ws.cell(row=idx, column=1).fill = self.coverage_fills['None']
                ws.cell(row=idx, column=2).fill = self.coverage_fills['None']

        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        for col_letter, width in self.rtm_column_widths.items():
            ws.column_dimensions[col_letter].width = width

        # ================================================================
        # FREEZE HEADER ROW
        # ================================================================
        ws.freeze_panes = f'A{header_row + 1}'

        # ================================================================
        # AUTO-FILTER
        # ================================================================
        if matrix:
            filter_range = f'A{header_row}:G{header_row + len(matrix)}'
            ws.auto_filter.ref = filter_range

    def _adjust_row_heights(self, ws, min_height: int = 15, max_height: int = 100):
        """
        PURPOSE:
            Adjust row heights based on content.

        PARAMETERS:
            ws: The worksheet
            min_height (int): Minimum row height
            max_height (int): Maximum row height (prevents huge rows)

        WHY THIS APPROACH:
            Text wrapping requires adequate row height to display
            all content. We estimate based on newline count.

        NOTE:
            This is a basic implementation. openpyxl doesn't have
            true auto-height like Excel's GUI.
        """
        for row in ws.iter_rows(min_row=2):
            max_lines = 1
            for cell in row:
                if cell.value:
                    # Count newlines to estimate needed height
                    lines = str(cell.value).count('\n') + 1
                    max_lines = max(max_lines, lines)

            # Calculate height: ~15 points per line
            calculated_height = min(max_lines * 15, max_height)
            calculated_height = max(calculated_height, min_height)
            ws.row_dimensions[row[0].row].height = calculated_height


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================
# This function provides a quick way to export without creating a class instance.
# Usage: from formatters.excel_formatter import export_to_excel
# ============================================================================

def export_to_excel(
    user_stories: list[dict],
    test_cases: list[dict],
    output_dir: str = "outputs/excel",
    source_file: Optional[str] = None,
    filename: str = "uat_package.xlsx",
    traceability_matrix: dict = None
) -> str:
    """
    PURPOSE:
        Convenience function to export user stories and tests to Excel.
        Creates a formatter, runs it, and returns the filepath.

    R EQUIVALENT:
        # In R with openxlsx:
        # export_to_excel <- function(stories, tests, output_dir = "outputs/excel",
        #                             source_file = NULL, filename = "uat_package.xlsx") {
        #     wb <- createWorkbook()
        #     # ... create sheets ...
        #     saveWorkbook(wb, file.path(output_dir, filename))
        # }

    PARAMETERS:
        user_stories (list[dict]): User story dicts from UserStoryGenerator
        test_cases (list[dict]): Test case dicts from UATGenerator
        output_dir (str): Where to save the file. Default: "outputs/excel"
        source_file (str): Name of source file for documentation. Default: None
        filename (str): Output filename. Default: "uat_package.xlsx"
        traceability_matrix (dict, optional): RTM from TraceabilityGenerator.
            If provided, adds a Traceability Matrix sheet.

    RETURNS:
        str: Path to created Excel file

    EXAMPLE:
        from formatters.excel_formatter import export_to_excel

        filepath = export_to_excel(
            stories,
            test_cases,
            source_file="requirements.xlsx",
            filename="project_uat.xlsx"
        )
        print(f"Created: {filepath}")
    """
    formatter = ExcelFormatter(
        output_dir=output_dir,
        source_file=source_file
    )

    return formatter.export(
        user_stories=user_stories,
        test_cases=test_cases,
        filename=filename,
        traceability_matrix=traceability_matrix
    )


# ============================================================================
# STANDALONE TEST
# ============================================================================
# Run this file directly to test with sample data:
#     python formatters/excel_formatter.py
# ============================================================================

if __name__ == "__main__":
    # Sample data for testing
    sample_stories = [
        {
            'generated_id': 'US-001',
            'title': 'User login with email and password',
            'user_story': 'As a user, I want to login with my email and password, so that I can access my account securely.',
            'priority': 'High',
            'role': 'user',
            'acceptance_criteria': [
                'Given valid credentials, user is logged in successfully',
                'Given invalid credentials, error message is displayed',
                'Session expires after 30 minutes of inactivity'
            ],
            'flags': [],
            'source_requirement': {'row_number': 5}
        },
        {
            'generated_id': 'US-002',
            'title': 'Export dashboard data to Excel',
            'user_story': 'As an analyst, I want to export dashboard data to Excel, so that I can perform offline analysis.',
            'priority': 'Medium',
            'role': 'analyst',
            'acceptance_criteria': [
                'Export button visible on dashboard',
                'Excel file downloads within 10 seconds',
                'All visible data included in export'
            ],
            'flags': ['priority_inferred'],
            'source_requirement': {'row_number': 12}
        }
    ]

    sample_tests = [
        {
            'test_id': 'TEST-AUTH-001',
            'source_story_id': 'US-001',
            'category': 'Authentication',
            'title': 'Verify successful login with valid credentials',
            'test_type': 'happy_path',
            'prerequisites': ['Valid test user exists', 'User is not logged in'],
            'test_steps': [
                '1. Navigate to login page',
                '2. Enter valid email',
                '3. Enter valid password',
                '4. Click Login'
            ],
            'expected_results': [
                '• User is redirected to dashboard',
                '• Welcome message displayed',
                '• User name appears in header'
            ],
            'moscow': 'Must Have',
            'est_time': '5 min',
            'notes': 'Test on Chrome, Firefox, Safari'
        },
        {
            'test_id': 'TEST-AUTH-002',
            'source_story_id': 'US-001',
            'category': 'Authentication',
            'title': 'Verify login fails with invalid password',
            'test_type': 'negative',
            'prerequisites': ['Valid test user exists'],
            'test_steps': [
                '1. Navigate to login page',
                '2. Enter valid email',
                '3. Enter wrong password',
                '4. Click Login'
            ],
            'expected_results': [
                '• Error message displayed',
                '• User remains on login page',
                '• Failed attempt is logged'
            ],
            'moscow': 'Must Have',
            'est_time': '3 min',
            'notes': 'Do not reveal which field is wrong'
        },
        {
            'test_id': 'TEST-RPT-001',
            'source_story_id': 'US-002',
            'category': 'Reporting',
            'title': 'Verify Excel export downloads successfully',
            'test_type': 'happy_path',
            'prerequisites': ['User is logged in', 'Dashboard has data'],
            'test_steps': [
                '1. Navigate to dashboard',
                '2. Click Export button',
                '3. Wait for download'
            ],
            'expected_results': [
                '• Excel file downloads',
                '• File opens correctly',
                '• Data matches dashboard'
            ],
            'moscow': 'Should Have',
            'est_time': '5 min',
            'notes': 'Check file size is reasonable'
        },
        {
            'test_id': 'TEST-RPT-002',
            'source_story_id': 'US-002',
            'category': 'Reporting',
            'title': 'Verify export handles empty data gracefully',
            'test_type': 'edge_case',
            'prerequisites': ['User is logged in', 'Dashboard has no data'],
            'test_steps': [
                '1. Navigate to empty dashboard',
                '2. Click Export button',
                '3. Observe behavior'
            ],
            'expected_results': [
                '• Informative message shown',
                '• No empty file generated'
            ],
            'moscow': 'Could Have',
            'est_time': '3 min',
            'notes': 'Edge case for empty state'
        }
    ]

    print("=" * 70)
    print("EXCEL FORMATTER TEST")
    print("=" * 70)
    print()

    # Export using the convenience function
    filepath = export_to_excel(
        sample_stories,
        sample_tests,
        source_file="test_requirements.xlsx",
        filename="test_uat_package.xlsx"
    )

    print(f"Created: {filepath}")
    print()

    # Verify file exists and show size
    import os
    if os.path.exists(filepath):
        size_kb = os.path.getsize(filepath) / 1024
        print(f"File size: {size_kb:.1f} KB")
        print()
        print("Open the file in Excel to verify formatting:")
        print(f"  open '{filepath}'")
    else:
        print("ERROR: File was not created!")
