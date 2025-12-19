# formatters/draft_excel_formatter.py
# ============================================================================
# PURPOSE: Create Excel file for human review of draft user stories
#
# This formatter creates an Excel workbook optimized for stakeholder review:
#   - Editable columns for refinement (title, story, acceptance criteria)
#   - Locked columns for reference (source requirement, row number)
#   - Status dropdown for approval workflow
#   - Instructions sheet for reviewers
#
# TWO-PHASE WORKFLOW:
#   Phase 1 (Draft): Generate stories → Export with this formatter → Human review
#   Phase 2 (Final): Import refined stories → Generate UAT test cases
#
# AVIATION ANALOGY:
#   Like creating a draft flight plan that needs pilot review and sign-off
#   before it becomes the official plan. The draft has all the computed data,
#   but the pilot adds their experience and local knowledge.
#
# ============================================================================

import os
from datetime import datetime
from typing import Optional

# openpyxl for Excel file creation
try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, Protection
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:
    raise ImportError("openpyxl required. Install with: pip3 install openpyxl")


class DraftExcelFormatter:
    """
    PURPOSE:
        Create an Excel file for human review of draft user stories.

    KEY FEATURES:
        - Editable columns (white background) for refinement
        - Locked columns (grey background) for reference only
        - Status dropdown: Draft, Pending Client Review, Approved, Needs Discussion, Out of Scope
        - Priority dropdown: Critical, High, Medium, Low
        - Is Technical dropdown: Yes, No
        - Client Feedback column for client signoff comments
        - Instructions sheet with client review workflow
        - Color-coded status column (Yellow, Blue, Green, Orange, Grey)

    CLIENT REVIEW WORKFLOW:
        1. Internal team reviews and refines draft stories
        2. Set status to "Pending Client Review" and send to client
        3. Client adds feedback in "Client Feedback" column
        4. Client changes status to "Approved" or "Needs Discussion"
        5. Once approved, run --phase final to generate UAT

    OUTPUT:
        Excel file with two sheets:
        1. "User Stories for Review" - Main review sheet (14 columns)
        2. "Instructions" - How to use the review sheet

    USAGE:
        formatter = DraftExcelFormatter(prefix="PROP")
        filepath = formatter.export(stories, output_dir="outputs/drafts")
    """

    # ========================================================================
    # STYLE DEFINITIONS
    # ========================================================================

    # Header styles
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data styles
    DATA_ALIGNMENT = Alignment(horizontal="left", vertical="top", wrap_text=True)
    DATA_ALIGNMENT_CENTER = Alignment(horizontal="center", vertical="top")

    # Editable vs locked styles
    EDITABLE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    LOCKED_FILL = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    # Status colors
    STATUS_FILLS = {
        'Draft': PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid"),  # Yellow
        'Approved': PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),  # Green
        'Needs Discussion': PatternFill(start_color="FFB366", end_color="FFB366", fill_type="solid"),  # Orange
        'Out of Scope': PatternFill(start_color="C0C0C0", end_color="C0C0C0", fill_type="solid"),  # Grey
        'Pending Client Review': PatternFill(start_color="87CEEB", end_color="87CEEB", fill_type="solid"),  # Light Blue
    }

    # Border
    CELL_BORDER = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Column widths
    COLUMN_WIDTHS = {
        'A': 18,   # Story ID (locked)
        'B': 30,   # Title (editable)
        'C': 50,   # User Story (editable)
        'D': 15,   # Role (editable)
        'E': 60,   # Acceptance Criteria (editable)
        'F': 40,   # Success Metrics (editable)
        'G': 12,   # Priority (dropdown)
        'H': 22,   # Status (dropdown) - wider for "Pending Client Review"
        'I': 30,   # Your Notes (editable)
        'J': 30,   # Meeting Context (editable)
        'K': 35,   # Client Feedback (editable) - NEW
        'L': 12,   # Is Technical (dropdown)
        'M': 50,   # Source Requirement (locked)
        'N': 10,   # Source Row (locked)
    }

    def __init__(
        self,
        prefix: str = "REQ",
        source_filename: Optional[str] = None
    ):
        """
        PURPOSE:
            Initialize the draft formatter.

        PARAMETERS:
            prefix (str): Story ID prefix (e.g., "PROP", "GRX")
            source_filename (str): Original input filename for reference
        """
        self.prefix = prefix
        self.source_filename = source_filename or "Unknown"

    def export(
        self,
        stories: list[dict],
        output_dir: str = "outputs/drafts"
    ) -> str:
        """
        PURPOSE:
            Export user stories to Excel for human review.

        PARAMETERS:
            stories (list[dict]): User stories from UserStoryGenerator
            output_dir (str): Output directory

        RETURNS:
            str: Path to created Excel file
        """
        # Ensure output directory exists
        os.makedirs(output_dir, exist_ok=True)

        # Create workbook
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        # Create sheets
        self._create_review_sheet(wb, stories)
        self._create_instructions_sheet(wb)

        # Generate filename
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")
        filename = f"{self.prefix}_draft_user_stories_{timestamp}.xlsx"
        filepath = os.path.join(output_dir, filename)

        # Save workbook
        wb.save(filepath)

        return filepath

    def _create_review_sheet(self, wb: Workbook, stories: list[dict]):
        """
        PURPOSE:
            Create the main review sheet with user stories.

        COLUMNS:
            A: Story ID (locked)
            B: Title (editable)
            C: User Story (editable)
            D: Role (editable)
            E: Acceptance Criteria (editable)
            F: Success Metrics (editable)
            G: Priority (dropdown)
            H: Status (dropdown)
            I: Your Notes (editable) - internal team comments
            J: Meeting Context (editable) - decisions from meetings
            K: Client Feedback (editable) - client signoff comments
            L: Is Technical (dropdown)
            M: Source Requirement (locked)
            N: Source Row (locked)
        """
        ws = wb.create_sheet("User Stories for Review")

        # ================================================================
        # HEADER ROW
        # ================================================================
        headers = [
            ('Story ID', False),           # A - locked
            ('Title', True),               # B - editable
            ('User Story', True),          # C - editable
            ('Role', True),                # D - editable
            ('Acceptance Criteria', True), # E - editable
            ('Success Metrics', True),     # F - editable
            ('Priority', True),            # G - dropdown
            ('Status', True),              # H - dropdown
            ('Your Notes', True),          # I - editable (internal)
            ('Meeting Context', True),     # J - editable
            ('Client Feedback', True),     # K - editable (client signoff)
            ('Is Technical', True),        # L - dropdown
            ('Source Requirement', False), # M - locked
            ('Source Row', False),         # N - locked
        ]

        for col, (header, _) in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.alignment = self.HEADER_ALIGNMENT
            cell.border = self.CELL_BORDER

        # ================================================================
        # DATA ROWS
        # ================================================================
        for row_idx, story in enumerate(stories, 2):
            # A: Story ID (locked)
            cell = ws.cell(row=row_idx, column=1, value=story.get('generated_id', 'N/A'))
            cell.fill = self.LOCKED_FILL
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # B: Title (editable)
            cell = ws.cell(row=row_idx, column=2, value=story.get('title', ''))
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # C: User Story (editable)
            cell = ws.cell(row=row_idx, column=3, value=story.get('user_story', ''))
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # D: Role (editable)
            cell = ws.cell(row=row_idx, column=4, value=story.get('role', 'user'))
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # E: Acceptance Criteria (editable) - join list with newlines
            ac = story.get('acceptance_criteria', [])
            ac_text = '\n'.join(ac) if isinstance(ac, list) else str(ac)
            cell = ws.cell(row=row_idx, column=5, value=ac_text)
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # F: Success Metrics (editable) - extract from acceptance criteria
            # Success metrics are typically after "SUCCESS METRICS:" header
            metrics_text = self._extract_success_metrics(ac)
            cell = ws.cell(row=row_idx, column=6, value=metrics_text)
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # G: Priority (dropdown)
            cell = ws.cell(row=row_idx, column=7, value=story.get('priority', 'Medium'))
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # H: Status (dropdown) - default to Draft
            status = 'Draft'
            # Check if story was flagged as out_of_scope or completed
            flags = story.get('flags', [])
            if 'out_of_scope' in flags:
                status = 'Out of Scope'
            elif 'completed' in flags:
                status = 'Approved'
            elif 'needs_clarification' in flags or 'vague_language' in flags:
                status = 'Needs Discussion'

            cell = ws.cell(row=row_idx, column=8, value=status)
            cell.fill = self.STATUS_FILLS.get(status, self.EDITABLE_FILL)
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # I: Your Notes (editable) - empty for internal reviewer
            cell = ws.cell(row=row_idx, column=9, value='')
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # J: Meeting Context (editable) - empty for reviewer
            cell = ws.cell(row=row_idx, column=10, value='')
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # K: Client Feedback (editable) - empty for client signoff
            cell = ws.cell(row=row_idx, column=11, value='')
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # L: Is Technical (dropdown)
            is_tech = 'Yes' if story.get('is_technical', True) else 'No'
            cell = ws.cell(row=row_idx, column=12, value=is_tech)
            cell.fill = self.EDITABLE_FILL
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # M: Source Requirement (locked)
            source_req = story.get('source_requirement', {})
            source_text = story.get('description', '')[:500]  # Truncate if too long
            cell = ws.cell(row=row_idx, column=13, value=source_text)
            cell.fill = self.LOCKED_FILL
            cell.alignment = self.DATA_ALIGNMENT

            # N: Source Row (locked)
            source_row = source_req.get('row_number', 'N/A')
            cell = ws.cell(row=row_idx, column=14, value=source_row)
            cell.fill = self.LOCKED_FILL
            cell.alignment = self.DATA_ALIGNMENT_CENTER

            # Apply border to all cells (now 14 columns)
            for col in range(1, 15):
                ws.cell(row=row_idx, column=col).border = self.CELL_BORDER

        # ================================================================
        # DATA VALIDATION (Dropdowns)
        # ================================================================
        max_row = len(stories) + 1

        # Priority dropdown (column G)
        priority_dv = DataValidation(
            type="list",
            formula1='"Critical,High,Medium,Low"',
            allow_blank=False
        )
        priority_dv.error = "Please select a valid priority"
        priority_dv.errorTitle = "Invalid Priority"
        ws.add_data_validation(priority_dv)
        priority_dv.add(f'G2:G{max_row}')

        # Status dropdown (column H) - includes Pending Client Review
        status_dv = DataValidation(
            type="list",
            formula1='"Draft,Approved,Needs Discussion,Out of Scope,Pending Client Review"',
            allow_blank=False
        )
        status_dv.error = "Please select a valid status"
        status_dv.errorTitle = "Invalid Status"
        ws.add_data_validation(status_dv)
        status_dv.add(f'H2:H{max_row}')

        # Is Technical dropdown (column L - shifted from K)
        tech_dv = DataValidation(
            type="list",
            formula1='"Yes,No"',
            allow_blank=False
        )
        tech_dv.error = "Please select Yes or No"
        tech_dv.errorTitle = "Invalid Value"
        ws.add_data_validation(tech_dv)
        tech_dv.add(f'L2:L{max_row}')

        # ================================================================
        # COLUMN WIDTHS
        # ================================================================
        for col_letter, width in self.COLUMN_WIDTHS.items():
            ws.column_dimensions[col_letter].width = width

        # ================================================================
        # FREEZE HEADER ROW
        # ================================================================
        ws.freeze_panes = 'A2'

        # ================================================================
        # AUTO-FILTER
        # ================================================================
        ws.auto_filter.ref = ws.dimensions

    def _extract_success_metrics(self, acceptance_criteria: list) -> str:
        """Extract success metrics from acceptance criteria list."""
        metrics = []
        in_metrics = False

        for line in acceptance_criteria:
            if 'SUCCESS METRICS' in line:
                in_metrics = True
                continue
            elif 'ACCEPTANCE CRITERIA' in line or 'PROCESS CHANGE' in line:
                in_metrics = False
                continue

            if in_metrics and line.strip().startswith('•'):
                metrics.append(line.strip())

        return '\n'.join(metrics) if metrics else ''

    def _create_instructions_sheet(self, wb: Workbook):
        """
        PURPOSE:
            Create instructions sheet for reviewers.
        """
        ws = wb.create_sheet("Instructions")

        instructions = [
            ("How to Use This Review Sheet", ""),
            ("", ""),
            ("1. INTERNAL REVIEW PROCESS", ""),
            ("   • Review each user story in the 'User Stories for Review' sheet", ""),
            ("   • Refine the Title, User Story, and Acceptance Criteria as needed", ""),
            ("   • Add context from meetings in the 'Meeting Context' column", ""),
            ("   • Add internal notes in the 'Your Notes' column", ""),
            ("   • Set Status to 'Pending Client Review' when ready for client", ""),
            ("", ""),
            ("2. CLIENT REVIEW WORKFLOW", ""),
            ("   Step 1: Internal team reviews and refines draft stories", ""),
            ("   Step 2: Set status to 'Pending Client Review' and send to client", ""),
            ("   Step 3: Client adds feedback in 'Client Feedback' column", ""),
            ("   Step 4: Client changes status to 'Approved' or 'Needs Discussion'", ""),
            ("   Step 5: Once approved, run --phase final to generate UAT", ""),
            ("", ""),
            ("3. STATUS OPTIONS", ""),
            ("   • Draft - Not yet reviewed, needs internal attention", "Yellow"),
            ("   • Pending Client Review - Sent to client, awaiting response", "Blue"),
            ("   • Approved - Client approved, ready for UAT generation", "Green"),
            ("   • Needs Discussion - Has questions or concerns to resolve", "Orange"),
            ("   • Out of Scope - Removed from current scope, will be skipped", "Grey"),
            ("", ""),
            ("4. COLUMN GUIDE", ""),
            ("   Story ID - Auto-generated, do not modify", "Locked"),
            ("   Title - Edit to make more descriptive", "Editable"),
            ("   User Story - Refine the 'As a... I want... so that...' format", "Editable"),
            ("   Role - The user role (analyst, coordinator, etc.)", "Editable"),
            ("   Acceptance Criteria - What must be true for this story to be 'done'", "Editable"),
            ("   Success Metrics - Measurable outcomes", "Editable"),
            ("   Priority - Critical, High, Medium, Low", "Dropdown"),
            ("   Status - See status options above", "Dropdown"),
            ("   Your Notes - Internal team comments and questions", "Editable"),
            ("   Meeting Context - Decisions from stakeholder meetings", "Editable"),
            ("   Client Feedback - Client signoff comments and responses", "Editable"),
            ("   Is Technical - Yes for features, No for workflow/process changes", "Dropdown"),
            ("   Source Requirement - Original text for reference", "Locked"),
            ("   Source Row - Row number from original file", "Locked"),
            ("", ""),
            ("5. GENERATING UAT TEST CASES", ""),
            ("   After client approval, save this file and run:", ""),
            (f"   python3 run.py \"this_file.xlsx\" --prefix {self.prefix} --phase final", ""),
            ("", ""),
            ("   This will generate UAT test cases from your refined stories.", ""),
            ("   Stories with Status = 'Out of Scope' will be skipped.", ""),
            ("   Stories with Status = 'Needs Discussion' will be flagged.", ""),
            ("   Stories with Status = 'Pending Client Review' will be included.", ""),
        ]

        # Title style
        title_font = Font(bold=True, size=14, color="4472C4")
        section_font = Font(bold=True, size=11)
        normal_font = Font(size=10)

        for row_idx, (text, note) in enumerate(instructions, 1):
            cell = ws.cell(row=row_idx, column=1, value=text)

            if row_idx == 1:
                cell.font = title_font
            elif text.endswith(':') or (text.strip() and text[0].isdigit()):
                cell.font = section_font
            else:
                cell.font = normal_font

            if note:
                note_cell = ws.cell(row=row_idx, column=2, value=note)
                note_cell.font = Font(size=10, italic=True, color="666666")

        # Column widths
        ws.column_dimensions['A'].width = 70
        ws.column_dimensions['B'].width = 20


# ============================================================================
# CONVENIENCE FUNCTION
# ============================================================================

def export_draft_for_review(
    stories: list[dict],
    prefix: str = "REQ",
    output_dir: str = "outputs/drafts",
    source_filename: Optional[str] = None
) -> str:
    """
    PURPOSE:
        Convenience function to export draft stories for review.

    PARAMETERS:
        stories (list[dict]): User stories from UserStoryGenerator
        prefix (str): Story ID prefix
        output_dir (str): Output directory
        source_filename (str): Original input filename

    RETURNS:
        str: Path to created Excel file
    """
    formatter = DraftExcelFormatter(prefix=prefix, source_filename=source_filename)
    return formatter.export(stories, output_dir=output_dir)


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("DRAFT EXCEL FORMATTER TEST")
    print("=" * 70)

    # Sample stories
    sample_stories = [
        {
            'generated_id': 'TEST-RECRUIT-001',
            'title': 'Number (N) Invited',
            'user_story': 'As a research coordinator, I want to track number of patients invited, so that I can measure outreach effectiveness',
            'role': 'research coordinator',
            'priority': 'High',
            'is_technical': True,
            'acceptance_criteria': [
                'ACCEPTANCE CRITERIA:',
                '• Count of invited patients displays accurately',
                '• Metrics are segmented by program',
                '',
                'SUCCESS METRICS:',
                '• Zero discrepancy between source and display',
            ],
            'description': 'Track number (N) of patients who received invitations',
            'flags': [],
            'source_requirement': {'row_number': 2},
        },
        {
            'generated_id': 'TEST-WF-001',
            'title': 'Manual Review Process',
            'user_story': '[WORKFLOW CHANGE] Update the manual review workflow',
            'role': 'staff',
            'priority': 'Low',
            'is_technical': False,
            'acceptance_criteria': [
                'PROCESS CHANGE REQUIREMENTS:',
                '• Process documentation is updated',
            ],
            'description': 'Update manual review workflow',
            'flags': ['workflow_change'],
            'source_requirement': {'row_number': 5},
        },
    ]

    formatter = DraftExcelFormatter(prefix="TEST")
    filepath = formatter.export(sample_stories, output_dir="outputs/drafts")

    print(f"\n✓ Created draft review file: {filepath}")
    print("\nInstructions:")
    print("  1. Open the file in Excel")
    print("  2. Review and refine the stories")
    print("  3. Set Status for each story")
    print("  4. Run: python3 run.py \"<filepath>\" --prefix TEST --phase final")
