# formatters/audit_excel_formatter.py
# ============================================================================
# AUDIT HISTORY EXCEL FORMATTER
# ============================================================================
# Purpose: Export audit history data to Excel for compliance documentation
#          and regulatory reviews.
#
# AVIATION ANALOGY:
#   This is like generating the official flight data transcript from the
#   black box - formatted, timestamped, and suitable for regulatory review.
#
# REGULATORY CONTEXT:
#   FDA 21 CFR Part 11 requires "the ability to generate accurate and
#   complete copies of records in both human readable and electronic form."
#   This formatter creates human-readable Excel reports for compliance audits.
#
# ============================================================================

import os
from datetime import datetime
from typing import Optional, Dict, List, Any

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# ============================================================================
# COLOR SCHEME FOR ACTIONS
# ============================================================================
# Color coding helps quickly identify types of changes

ACTION_COLORS = {
    'Created': 'C6EFCE',      # Light green
    'Updated': 'BDD7EE',      # Light blue
    'Deleted': 'FFC7CE',      # Light red
    'Status Changed': 'FFEB9C',  # Light yellow/gold
    'Approved': '92D050',     # Bright green
    'Imported': 'E2EFDA',     # Pale green
    'Exported': 'DDEBF7',     # Pale blue
    'Test Executed': 'FCE4D6',  # Light orange
}

# Record type colors for the Type column
RECORD_TYPE_COLORS = {
    'client': 'D9E1F2',       # Light blue-gray
    'program': 'D9E1F2',      # Light blue-gray
    'requirement': 'FFF2CC',  # Light yellow
    'user_story': 'E2EFDA',   # Light green
    'test_case': 'DDEBF7',    # Light blue
    'compliance_gap': 'FCE4D6',  # Light orange
    'traceability': 'F2F2F2', # Light gray
}


def export_audit_to_excel(
    audit_entries: List[Dict],
    output_path: str,
    report_title: str = "Audit Report",
    program_info: Optional[Dict] = None,
    date_range: Optional[Dict] = None,
    summary: Optional[Dict] = None
) -> str:
    """
    PURPOSE:
        Export audit history to Excel for compliance documentation.
        Creates a professionally formatted workbook suitable for
        regulatory reviews.

    PARAMETERS:
        audit_entries (list): List of audit entry dicts from database
        output_path (str): Path to save the Excel file
        report_title (str): Title for the report header
        program_info (dict, optional): Program details to include
        date_range (dict, optional): {'start': '...', 'end': '...'}
        summary (dict, optional): Summary statistics

    RETURNS:
        str: Path to the created Excel file

    OUTPUT STRUCTURE:
        Sheet 1: "Audit Trail"
            - Header with report title, date range, generation timestamp
            - Summary statistics (if provided)
            - Detailed audit entries table

    COLUMNS:
        - Date/Time (formatted as datetime)
        - Record Type (color coded)
        - Record ID
        - Action (color coded)
        - Field Changed
        - Old Value
        - New Value
        - Changed By
        - Reason

    FORMATTING:
        - Header row frozen and bold
        - Date column formatted as datetime
        - Auto-filter enabled
        - Column widths auto-sized
        - Color coding for actions and record types

    R EQUIVALENT:
        # In R, you'd use openxlsx package:
        # wb <- createWorkbook()
        # addWorksheet(wb, "Audit Trail")
        # writeData(wb, "Audit Trail", audit_data)
        # addStyle(wb, "Audit Trail", headerStyle, rows = 1)

    RAISES:
        ImportError: If openpyxl is not installed
        IOError: If file cannot be written
    """
    # ========================================================================
    # VALIDATION
    # ========================================================================
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install it with: pip install openpyxl"
        )

    # Ensure output directory exists
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)

    # ========================================================================
    # CREATE WORKBOOK
    # ========================================================================
    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Trail"

    # ========================================================================
    # STYLES
    # ========================================================================
    # Title style
    title_font = Font(name='Calibri', size=16, bold=True, color='1F4E79')

    # Header style (for column headers)
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4',
                              fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center',
                                 wrap_text=True)

    # Metadata style
    meta_font = Font(name='Calibri', size=10, italic=True, color='666666')

    # Border style
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )

    # ========================================================================
    # HEADER SECTION
    # ========================================================================
    current_row = 1

    # Title
    ws.cell(row=current_row, column=1, value=report_title)
    ws.cell(row=current_row, column=1).font = title_font
    ws.merge_cells(start_row=current_row, start_column=1,
                   end_row=current_row, end_column=5)
    current_row += 1

    # Generation timestamp
    timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    ws.cell(row=current_row, column=1, value=f"Generated: {timestamp}")
    ws.cell(row=current_row, column=1).font = meta_font
    current_row += 1

    # Date range (if provided)
    if date_range:
        start = date_range.get('start', 'N/A')
        end = date_range.get('end', 'N/A')
        ws.cell(row=current_row, column=1, value=f"Date Range: {start} to {end}")
        ws.cell(row=current_row, column=1).font = meta_font
        current_row += 1

    # Program info (if provided)
    if program_info:
        prog_name = program_info.get('name', 'N/A')
        prefix = program_info.get('prefix', 'N/A')
        ws.cell(row=current_row, column=1, value=f"Program: {prog_name} ({prefix})")
        ws.cell(row=current_row, column=1).font = meta_font
        current_row += 1

    # Blank row
    current_row += 1

    # ========================================================================
    # SUMMARY SECTION (if provided)
    # ========================================================================
    if summary and summary.get('total_changes', 0) > 0:
        # Summary header
        ws.cell(row=current_row, column=1, value="Summary")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
        current_row += 1

        # Total changes
        ws.cell(row=current_row, column=1, value=f"Total Changes: {summary.get('total_changes', 0)}")
        current_row += 1

        # By action
        by_action = summary.get('by_action', {})
        if by_action:
            actions_str = ", ".join([f"{k}: {v}" for k, v in sorted(by_action.items())])
            ws.cell(row=current_row, column=1, value=f"By Action: {actions_str}")
            current_row += 1

        # By record type
        by_type = summary.get('by_type', {})
        if by_type:
            types_str = ", ".join([f"{k}: {v}" for k, v in sorted(by_type.items())])
            ws.cell(row=current_row, column=1, value=f"By Type: {types_str}")
            current_row += 1

        # Blank row before data
        current_row += 1

    # ========================================================================
    # DATA TABLE HEADERS
    # ========================================================================
    headers = [
        'Date/Time',
        'Record Type',
        'Record ID',
        'Action',
        'Field Changed',
        'Old Value',
        'New Value',
        'Changed By',
        'Reason'
    ]

    header_row = current_row

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    current_row += 1

    # ========================================================================
    # DATA ROWS
    # ========================================================================
    for entry in audit_entries:
        row_data = [
            entry.get('changed_date', ''),
            entry.get('record_type', ''),
            entry.get('record_id', ''),
            entry.get('action', ''),
            entry.get('field_changed', ''),
            entry.get('old_value', ''),
            entry.get('new_value', ''),
            entry.get('changed_by', ''),
            entry.get('change_reason', '')
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='top', wrap_text=True)

            # Apply color coding
            # Action column (column 4)
            if col == 4:
                action = str(value)
                if action in ACTION_COLORS:
                    cell.fill = PatternFill(
                        start_color=ACTION_COLORS[action],
                        end_color=ACTION_COLORS[action],
                        fill_type='solid'
                    )

            # Record Type column (column 2)
            elif col == 2:
                rec_type = str(value)
                if rec_type in RECORD_TYPE_COLORS:
                    cell.fill = PatternFill(
                        start_color=RECORD_TYPE_COLORS[rec_type],
                        end_color=RECORD_TYPE_COLORS[rec_type],
                        fill_type='solid'
                    )

        current_row += 1

    # ========================================================================
    # COLUMN WIDTHS
    # ========================================================================
    column_widths = {
        1: 20,   # Date/Time
        2: 14,   # Record Type
        3: 25,   # Record ID
        4: 16,   # Action
        5: 20,   # Field Changed
        6: 35,   # Old Value
        7: 35,   # New Value
        8: 12,   # Changed By
        9: 40,   # Reason
    }

    for col, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    # ========================================================================
    # FREEZE PANES AND AUTO-FILTER
    # ========================================================================
    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # Auto-filter on data range
    if audit_entries:
        last_row = header_row + len(audit_entries)
        ws.auto_filter.ref = f"A{header_row}:I{last_row}"

    # ========================================================================
    # SAVE WORKBOOK
    # ========================================================================
    wb.save(output_path)

    return output_path


def export_record_audit_trail(
    audit_entries: List[Dict],
    record_type: str,
    record_id: str,
    output_dir: str = "outputs/audit"
) -> str:
    """
    PURPOSE:
        Export audit trail for a single record to Excel.
        Convenience function with auto-generated filename.

    PARAMETERS:
        audit_entries (list): Audit entries for the record
        record_type (str): Type of record
        record_id (str): ID of the record
        output_dir (str): Directory to save the file

    RETURNS:
        str: Path to the created file
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    safe_id = record_id.replace('/', '_').replace('\\', '_')
    filename = f"audit_{record_type}_{safe_id}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    return export_audit_to_excel(
        audit_entries=audit_entries,
        output_path=output_path,
        report_title=f"Audit Trail: {record_id}"
    )


def export_program_audit_report(
    report_data: Dict,
    output_dir: str = "outputs/audit"
) -> str:
    """
    PURPOSE:
        Export program audit report to Excel.
        Convenience function with auto-generated filename.

    PARAMETERS:
        report_data (dict): Report data from get_program_audit_report()
        output_dir (str): Directory to save the file

    RETURNS:
        str: Path to the created file
    """
    program = report_data.get('program', {})
    prefix = program.get('prefix', 'UNKNOWN')
    date_range = report_data.get('date_range', {})

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"audit_report_{prefix}_{timestamp}.xlsx"
    output_path = os.path.join(output_dir, filename)

    return export_audit_to_excel(
        audit_entries=report_data.get('entries', []),
        output_path=output_path,
        report_title=f"Audit Report: {program.get('name', prefix)}",
        program_info=program,
        date_range=date_range,
        summary=report_data.get('summary')
    )


# ============================================================================
# STANDALONE TEST
# ============================================================================

if __name__ == "__main__":
    print("=" * 70)
    print("AUDIT EXCEL FORMATTER TEST")
    print("=" * 70)

    # Create sample audit entries
    sample_entries = [
        {
            'audit_id': 1,
            'record_type': 'user_story',
            'record_id': 'PROP-RECRUIT-001',
            'action': 'Created',
            'field_changed': None,
            'old_value': None,
            'new_value': 'Story: Number (N) Invited',
            'changed_by': 'system',
            'changed_date': '2024-12-19 10:30:00',
            'change_reason': 'Imported from Excel'
        },
        {
            'audit_id': 2,
            'record_type': 'user_story',
            'record_id': 'PROP-RECRUIT-001',
            'action': 'Updated',
            'field_changed': 'acceptance_criteria',
            'old_value': 'Original criteria...',
            'new_value': 'Updated criteria with new requirements...',
            'changed_by': 'system',
            'changed_date': '2024-12-19 11:00:00',
            'change_reason': None
        },
        {
            'audit_id': 3,
            'record_type': 'user_story',
            'record_id': 'PROP-RECRUIT-001',
            'action': 'Status Changed',
            'field_changed': 'status',
            'old_value': 'Draft',
            'new_value': 'Approved',
            'changed_by': 'admin',
            'changed_date': '2024-12-19 14:30:00',
            'change_reason': 'Client approved in review meeting'
        },
        {
            'audit_id': 4,
            'record_type': 'test_case',
            'record_id': 'PROP-RECRUIT-001-TC01',
            'action': 'Created',
            'field_changed': None,
            'old_value': None,
            'new_value': 'Test: Verify Number (N) Invited',
            'changed_by': 'system',
            'changed_date': '2024-12-19 15:00:00',
            'change_reason': 'Generated from user story'
        },
    ]

    # Export
    output_path = "outputs/audit/test_audit_report.xlsx"
    os.makedirs("outputs/audit", exist_ok=True)

    result = export_audit_to_excel(
        audit_entries=sample_entries,
        output_path=output_path,
        report_title="Test Audit Report",
        program_info={'name': 'Propel Analytics', 'prefix': 'PROP'},
        date_range={'start': '2024-12-01', 'end': '2024-12-31'},
        summary={
            'total_changes': 4,
            'by_type': {'user_story': 3, 'test_case': 1},
            'by_action': {'Created': 2, 'Updated': 1, 'Status Changed': 1}
        }
    )

    print(f"\n✓ Created: {result}")
    print("\nOpen the file to verify formatting.")
